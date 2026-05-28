"""
Універсальний апдейтер цін для ADU_Materials_Estimate.xlsx (V2 — phase-based).

Колонки на робочих листах V2 (17 шт):
A:1 #    B:2 Phase    C:3 Категорія    D:4 Item    E:5 Spec
F:6 Unit  G:7 Q-Low   H:8 Q-Mid       I:9 Q-High
J:10 Unit Price  K:11 Total (формула)
L:12 Source  M:13 SKU/Model  N:14 URL  O:15 Status  P:16 Confidence  Q:17 Notes
"""

from openpyxl import load_workbook
from datetime import date

XLSX = (
    "/Users/anatoliilisak/claude-workspace/documents/"
    "ADU — Menlo Park, CA 94025/ADU_Materials_Estimate.xlsx"
)

TODAY = date.today().isoformat()


def note_with_date(text):
    return f"{text} | Ціна збір: {TODAY}"


# === 18_Furniture_IKEA — Living Room (із попереднього збору) ===
# Структура порядку рядків per лист FURNITURE:
# Row 5: Sofa/Sleeper Sofa (Living)
# Row 6: Coffee Table (Living)
# Row 7: TV Bench (Living)
# Row 8: Storage Cabinet (Living)
# Row 9: Rug (Living)
# Row 10: Curtains + Blinds combo (Living)
# Row 11: Floor Lamp (Living)
# Rows 12-17: Bedroom items
# Rows 18-20: Dining items

FURNITURE_PRICES = [
    # (row, price, sku, url, confidence, status, notes_addendum)
    (
        5,
        899.00,
        "IKEA 405.512.31 — FRIHETEN",
        "https://www.ikea.com/us/en/p/friheten-sleeper-sofa-faringe-light-gray-40551231/",
        "high",
        "Confirmed",
        note_with_date("FRIHETEN sleeper Faringe light gray, 88⅝×41⅜. Queen bed 78×56. Alt: LYCKSELE LÖVÅS Ransta natural $499 (twin)"),
    ),
    (
        6,
        29.99,
        "IKEA 503.190.29 — LACK",
        "https://www.ikea.com/us/en/p/lack-coffee-table-white-stained-oak-effect-50319029/",
        "high",
        "Confirmed",
        note_with_date("LACK 35⅜×21⅝, white stained oak effect"),
    ),
    (
        7,
        75.00,
        "IKEA 005.760.02 — BESTÅ",
        "https://www.ikea.com/us/en/p/besta-tv-unit-white-stained-oak-effect-00576002/",
        "high",
        "Confirmed",
        note_with_date("BESTÅ 47¼\", white stained oak. До TV 45\". Якщо TV 55\" — обрати 70⅞\""),
    ),
    (
        8,
        44.99,
        "IKEA 603.245.20 — KALLAX 2x2",
        "https://www.ikea.com/us/en/p/kallax-shelf-unit-white-stained-oak-effect-60324520/",
        "high",
        "Confirmed",
        note_with_date("KALLAX 2x2 30⅛\", white stained oak"),
    ),
    (
        9,
        129.99,
        "IKEA 502.773.93 — LOHALS",
        "https://www.ikea.com/us/en/p/lohals-rug-flatwoven-natural-50277393/",
        "high",
        "Confirmed",
        note_with_date("LOHALS jute 5'3×7'7, natural"),
    ),
    (
        10,
        104.98,
        "IKEA 405.564.17 (RITVA) + 803.804.64 (TRETUR)",
        "https://www.ikea.com/us/en/p/ritva-curtains-with-tie-backs-1-pair-white-40556417/",
        "medium",
        "Confirmed",
        note_with_date("Combo: RITVA 57×84 pair $34.99 на French door + TRETUR roller 48×76¾ $69.99 на вікно"),
    ),
    (
        11,
        69.99,
        "IKEA 103.092.06 — VIDJA",
        "https://www.ikea.com/us/en/p/vidja-floor-lamp-white-10309206/",
        "high",
        "Confirmed",
        note_with_date("VIDJA white floor lamp"),
    ),
    # --- Bedroom (rows 12-17) ---
    (12, 329.00, "IKEA MALM Full з LÖNSET base",
     "https://www.ikea.com/us/en/p/malm-bed-frame-white-loenset-s39019169/",
     "high", "Confirmed", note_with_date("$329 з slatted base included")),
    (13, 250.00, "IKEA HÖVÅG/MORGEDAL Full mattress", "", "medium", "Confirmed", note_with_date("HAFSLO discontinued. HÖVÅG spring $299, MORGEDAL foam $229. Mid $250")),
    (14, 80.00, "IKEA MALM 2-drawer nightstand", "", "medium", "Confirmed", note_with_date("Per pc")),
    (15, 179.00, "IKEA MALM 4-drawer dresser", "", "medium", "Confirmed", note_with_date("")),
    (16, 30.00, "IKEA NYFORS/MELODI table lamp", "", "medium", "Confirmed", note_with_date("Per pc")),
    (17, 35.00, "IKEA RITVA curtain pair", "", "medium", "Confirmed", note_with_date("Same as Living combo")),
    # --- Dining (rows 18-20) ---
    (18, 130.00, "IKEA KRAGSTA round dining table 35\"", "", "medium", "Confirmed", note_with_date("DOCKSTA 41\" alt $179")),
    (19, 60.00, "IKEA TEODORES dining chair", "", "medium", "Confirmed", note_with_date("Per chair. JANINGE alt $79")),
    (20, 50.00, "IKEA FOTO pendant light (Optional)", "", "medium", "Optional", note_with_date("Qty 0 baseline")),
    # --- Bathroom storage (Optional) ---
    (21, 80.00, "IKEA bathroom storage cabinet + shelves", "", "medium", "Confirmed", note_with_date("Per §29.5. Mid підвищено 0→1 (review)")),
    # --- ДОДАНО per reviewer V2 ---
    (22, 25.00, "IKEA RÄCKA curtain rod + brackets set", "", "medium", "Confirmed", note_with_date("ДОДАНО (review)")),
    (23, 35.00, "IKEA RITVA panels for kitchen window", "", "medium", "Confirmed", note_with_date("ДОДАНО (review)")),
    (24, 35.00, "IKEA TUPPLUR roller blind for bathroom", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): privacy")),
    (25, 60.00, "Mattress protector + sheets/pillows/comforter allowance", "", "medium", "Confirmed", note_with_date("ДОДАНО (review)")),
    (26, 15.00, "Generic anti-tip strap kit", "", "medium", "Confirmed", note_with_date("ДОДАНО (review)")),
    (27, 10.00, "Generic felt pads / floor protectors", "", "medium", "Confirmed", note_with_date("ДОДАНО (review)")),
    (28, 400.00, "TV 43-55\" (Optional)", "", "medium", "Optional", note_with_date("ДОДАНО (review): qty 0 baseline")),
    (29, 150.00, "IKEA compact desk + chair (Optional)", "", "medium", "Optional", note_with_date("ДОДАНО (review)")),
    (30, 20.00, "Generic hangers set 50pc (Optional)", "", "medium", "Optional", note_with_date("ДОДАНО (review)")),
    (31, 40.00, "Bath mat + towels set (Optional)", "", "medium", "Optional", note_with_date("ДОДАНО (review)")),
]


# === 07_Electrical_Roughin — Phase 6 (повний) ===
# Структура (rows 5-35, 31 items): Panel + Wire + Devices + Lighting + Consumables + Bonding
ELECTRICAL_PRICES = [
    # --- PANEL ---
    (5, 139.00, "Square D QO120M100C 100A 20-space",
     "https://www.homedepot.com/p/Square-D-QO-100-Amp-20-Space-20-Circuit-Indoor-Main-Breaker-Load-Center-with-Cover-QO120M100C/100195842",
     "high", "Confirmed", note_with_date("$139 retail (promo $114 з HD card). 100A indoor main breaker load center")),
    (6, 130.00, "Square D QO standard breakers (non-AFCI only)", "", "medium", "Confirmed", note_with_date("Тільки для non-AFCI circuits: 240V (range/WH/HVAC) + bath GFCI + non-AFCI. AFCI breakers в r26 окремо")),
    # --- WIRE ---
    (7, 0.65, "Southwire 12/2 NM-B Romex SIMpull",
     "https://www.homedepot.com/p/Southwire-250-ft-12-2-Solid-Romex-SIMpull-CU-NM-B-W-G-Wire-28828269/202019375",
     "medium", "Confirmed", note_with_date("$0.65/LF (~$163 per 250ft roll). Bay Area premium")),
    (8, 0.45, "Southwire 14/2 NM-B Romex", "", "medium", "Confirmed", note_with_date("$0.45/LF lighter gauge")),
    (9, 2.50, "Southwire 8/3 + 10/2 NM-B (240V)", "", "medium", "Confirmed", note_with_date("Average $2.50/LF mixed. 8/3 для range, 10/2 для WH/HVAC")),
    (10, 0.80, "Allied EMT / Cantex PVC conduit 1/2-3/4\"", "", "medium", "TBD", note_with_date("Interior conduit per LF")),
    # --- DEVICES ---
    (11, 1.75, "Carlon plastic boxes mixed", "", "medium", "Confirmed", note_with_date("Average $1.75/box: 1g $1.50, 2g $2.50, ceiling $3, old-work $3.50, WP $12")),
    (12, 2.50, "Leviton TR-rated 15A/20A outlet", "", "medium", "Confirmed", note_with_date("Tamper-resistant per CA Code 2019+")),
    (13, 18.00, "Leviton/Eaton GFCI receptacle", "", "medium", "Confirmed", note_with_date("Per pc")),
    (14, 7.00, "Leviton switches + dimmers mix", "", "medium", "Confirmed", note_with_date("$2-4 single switch, $15-25 LED dimmer. Average $7")),
    (15, 0.75, "Leviton plastic cover plates", "", "medium", "Confirmed", note_with_date("Average per cover")),
    (16, 30.00, "Leviton bathroom fan timer switch", "", "medium", "TBD", note_with_date("Single-pole")),
    (17, 43.44, "First Alert 1039339 hardwired combo",
     "https://www.homedepot.com/p/First-Alert-Smoke-and-CO-Detector-AA-Battery-1039339/313143543",
     "medium", "Confirmed", note_with_date("Smoke+CO combo. CA Code: hardwired interconnected з battery backup")),
    (18, 22.00, "Leviton GFCI weatherproof + in-use cover", "", "medium", "TBD", note_with_date("Exterior")),
    (19, 40.00, "Square D 60A pull-out disconnect", "", "medium", "Confirmed", note_with_date("Mini-split outdoor disconnect")),
    (20, 15.00, "Leviton range/dryer/WH receptacle", "", "medium", "TBD", note_with_date("Mix 14-50R + 14-30R + 6-30R")),
    # --- LIGHTING FIXTURES ---
    (21, 20.00, "Halo/Lithonia LED wafer 4\"/6\" dimmable", "", "medium", "Confirmed", note_with_date("LED recessed, CA Title 24 compliant")),
    (22, 40.00, "Generic surface-mount LED ceiling", "", "medium", "TBD", note_with_date("Closet/utility")),
    (23, 60.00, "Generic LED wall mount exterior", "", "medium", "TBD", note_with_date("Entry + French door")),
    (24, 60.00, "Generic bathroom vanity LED", "", "medium", "TBD", note_with_date("Above mirror")),
    (25, 30.00, "Generic LED motion sensor light", "", "medium", "TBD", note_with_date("Closet")),
    # --- ДОДАНО V2 (CA Code + consumables) ---
    (26, 67.00, "Square D QO120DFC dual-function AFCI/GFCI",
     "https://www.homedepot.com/p/Square-D-QO-20-Amp-Single-Pole-Dual-Function-CAFCI-and-GFCI-Circuit-Breaker-6-Pack-QO120DFC6/206474144",
     "high", "Confirmed", note_with_date("$67/each (6-pack $401.31). ДОДАНО (CA Code 2019+): bedroom/living/kitchen/dining mandatory")),
    (27, 10.00, "Buchanan/Ideal wire nuts pack 50ct", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (28, 15.00, "3M Super 33 electrical tape pack 5pk", "", "medium", "Confirmed", note_with_date("ДОДАНО: black/white/red phasing")),
    (29, 12.00, "Gardner Bender Romex staples 1 lb", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (30, 15.00, "Generic Romex 3/8 connectors 25-pack", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (31, 10.00, "Generic #12 green ground pigtails pack", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (32, 2.00, "Carlon junction boxes plastic", "", "medium", "Confirmed", note_with_date("ДОДАНО: ceiling splices")),
    (33, 35.00, "Leviton T5832 GFCI+USB receptacle", "", "medium", "TBD", note_with_date("ДОДАНО: modern code allowance")),
    (34, 100.00, "Eaton CHSPT2ULTRA whole-house SPD type 2", "", "medium", "TBD", note_with_date("ДОДАНО: CA recommendation")),
    (35, 8.00, "Bridgeport bonding bushing + jumper", "", "medium", "Confirmed", note_with_date("ДОДАНО: metal water line bonding")),
    # --- ДОДАНО per reviewer V3 ---
    (36, 15.00, "Square D ground/neutral bar kit", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): окремо від subpanel")),
    (37, 8.00, "Generic panel/circuit label set", "", "medium", "Confirmed", note_with_date("ДОДАНО (review)")),
    (38, 35.00, "Mean Well/IKEA LED driver 60W", "", "medium", "Optional", note_with_date("ДОДАНО (review): under-cabinet LED — Optional, qty mid=1 only if обрано")),
    (39, 40.00, "Generic Cat6 + RG6 cable allowance", "", "medium", "Optional", note_with_date("ДОДАНО (review): low-voltage allowance — Optional")),
    (40, 40.00, "3M fire caulk + STI putty pads", "", "medium", "TBD", note_with_date("ДОДАНО (review)")),
    (41, 30.00, "Generic spare conduit + fittings", "", "medium", "Optional", note_with_date("ДОДАНО (review): future expansion — Optional")),
]


# === 02_Site_Exterior_Utilities — Phase 1 V3 (reviewer fixes + new items) ===
# Структура: Sewer rows 5-18 (14) + Water rows 19-33 (15) + Electrical rows 34-52 (19) + Site Prep row 53 (1)
SITE_EXTERIOR_PRICES = [
    # --- SEWER (rows 5-18) ---
    (5, 2.40, "Charlotte / JM Eagle 4\" SDR35",
     "https://www.homedepot.com/p/Charlotte-Pipe-4-in-x-10-ft-White-PVC-Sewer-Main-SDR35-Pipe-with-Gaskete-S-M-06004G-1400/203504420",
     "high", "Confirmed", note_with_date("$2.40/LF. Charlotte SDR35 10ft pieces ~$24")),
    (6, 5.50, "Generic SDR35 4\"", "", "medium", "Confirmed", note_with_date("Hub couplings")),
    (7, 7.50, "Generic SDR35 4\"", "", "medium", "Confirmed", note_with_date("45° + 90° short radius mix")),
    (8, 15.00, "Generic SDR35 4\"", "", "medium", "TBD", note_with_date("Long-radius sweep — slope verify by plumber")),
    (9, 15.00, "Generic SDR35 4\"", "", "medium", "Confirmed", note_with_date("Wye 4×4×4")),
    (10, 28.00, "Generic 4\" cleanout", "", "medium", "Confirmed", note_with_date("Body + plug + collar. Mid підвищено до 3")),
    (11, 12.00, "Oatey / Christy's", "", "medium", "Confirmed", note_with_date("Primer+cement combo")),
    (12, 35.00, "Aggregate yard", "", "medium", "Confirmed", note_with_date("Sand bedding — verify if required")),
    (13, 45.00, "Aggregate yard", "", "medium", "TBD", note_with_date("3/4\" rock backfill")),
    (14, 0.15, "Generic 14 AWG tracer", "", "medium", "Confirmed", note_with_date("Tracer wire — verify if required by code")),
    (15, 15.00, "Generic detectable", "", "medium", "Confirmed", note_with_date("'CAUTION SEWER'")),
    (16, 25.00, "4\" PVC sleeve", "", "low", "TBD", note_with_date("Slab penetration")),
    (17, 9.00, "Quikrete / Sakrete", "", "medium", "TBD", note_with_date("50 lb bag")),
    (18, 25.00, "Generic 4\" caps allowance", "", "medium", "Confirmed", note_with_date("ДОДАНО: test caps/plugs для pipe ends")),
    # --- WATER SUPPLY (rows 19-33) ---
    (19, 1.45, "SharkBite UA80W100 PEX-A 1\"",
     "https://www.homedepot.com/p/SharkBite-1-in-x-100-ft-Coil-White-PEX-A-Pipe-UA80W100/316358329",
     "high", "Confirmed", note_with_date("$1.45/LF ($145 per 100ft coil). Alt: copper L $4-5/LF")),
    (20, 1.00, "HDPE protective sleeve", "", "medium", "Confirmed", note_with_date("PEX-A UV protection for buried")),
    (21, 30.00, "Brass ball valve 1\"", "", "medium", "Confirmed", note_with_date("Full port, lead-free")),
    (22, 20.00, "Brass union 1\"", "", "medium", "Confirmed", note_with_date("Tap point + ADU connection")),
    (23, 250.00, "Badger / Neptune meter", "", "low", "Optional", note_with_date("OPTIONAL — verify з utility / town")),
    (24, 35.00, "PEX brass crimp pack", "", "medium", "Confirmed", note_with_date("Fittings assortment")),
    (25, 7.00, "Generic 1\" sleeve", "", "medium", "TBD", ""),
    (26, 0.70, "Armaflex foam 1\"", "", "medium", "TBD", note_with_date("Above-grade insulation")),
    (27, 0.15, "Generic 14 AWG tracer", "", "medium", "Confirmed", note_with_date("Tracer wire — verify if required by code")),
    (28, 12.00, "Generic 'WATER' tape", "", "medium", "Confirmed", ""),
    (29, 45.00, "Aggregate yard", "", "medium", "TBD", note_with_date("Sand bedding")),
    (30, 100.00, "Watts / Honeywell PRV", "", "low", "TBD", note_with_date("If street >80 PSI")),
    (31, 20.00, "Generic test cap set", "", "medium", "Confirmed", note_with_date("ДОДАНО: pressure test caps")),
    (32, 25.00, "Generic valve access box", "", "medium", "TBD", note_with_date("ДОДАНО: outdoor shut-off access")),
    (33, 20.00, "Generic pipe clamps pack", "", "medium", "Confirmed", note_with_date("ДОДАНО: above-grade pipe supports")),
    # --- ELECTRICAL FEED (rows 34-52) ---
    (34, 1.10, "Cantex / Carlon Sch40 2\"", "", "medium", "Confirmed", note_with_date("$11 per 10ft. Sch 40 PVC")),
    (35, 20.00, "Cantex / Carlon long-radius", "", "medium", "Confirmed", note_with_date("90° LR sweep для cable pulling")),
    (36, 15.00, "Cantex / Carlon", "", "medium", "Confirmed", note_with_date("45° sweep")),
    (37, 25.00, "Cantex expansion", "", "medium", "Confirmed", note_with_date("Thermal expansion — verify if req'd")),
    (38, 10.00, "Oatey PVC cement", "", "medium", "Confirmed", note_with_date("Conduit solvent weld")),
    (39, 8.78, "Cu SER 4-4-4-6 (base) / Al SER 2-2-2-4 (alt)",
     "https://www.wireandcableyourway.com/4-4-4-6-copper-ser-service-entrance-cable",
     "high", "TBD", note_with_date("TBD by electrician/load calc. Cu $8.78/LF, Al ~$3.50/LF (-$315)")),
    (40, 45.00, "Square D / Eaton 100A", "", "medium", "TBD", note_with_date("Non-fusible exterior if req'd")),
    (41, 18.00, "Copper-clad 8 ft rod", "", "medium", "Confirmed", note_with_date("Driven ground rod")),
    (42, 5.00, "Acorn ground clamp", "", "medium", "Confirmed", note_with_date("Bolt-on")),
    (43, 1.20, "#6 bare copper", "", "medium", "Confirmed", note_with_date("Per LF, service ground")),
    (44, 20.00, "Generic pull string / mule tape", "", "medium", "Confirmed", note_with_date("Cable pulling aid")),
    (45, 15.00, "NOALOX anti-oxidant", "", "medium", "TBD", note_with_date("Якщо AL feeder used")),
    (46, 20.00, "Carlon / Cantex straps pack", "", "medium", "Confirmed", note_with_date("ДОДАНО: mounting hardware")),
    (47, 15.00, "Generic adapters/bushings pack", "", "medium", "Confirmed", note_with_date("ДОДАНО: terminal adapters")),
    (48, 10.00, "Generic labels set", "", "medium", "Confirmed", note_with_date("ДОДАНО: subpanel/circuit labels")),
    (49, 10.00, "RACO / Carlon WP box", "", "medium", "Confirmed", ""),
    (50, 25.00, "Generic 6×6×4 PVC", "", "medium", "TBD", ""),
    (51, 12.00, "Generic 'ELECTRIC' red", "", "medium", "Confirmed", ""),
    (52, 25.00, "Adapters/bushings pack", "", "medium", "Confirmed", note_with_date("Stubs, male adapters, bell ends")),
    # --- SITE PREP (row 53) ---
    (53, 30.00, "Generic marking paint / stakes / string set", "", "medium", "Confirmed", note_with_date("ДОДАНО: trench layout aids")),
]


# === 03_Raised_Floor_System — Phase 2 ===
# Структура (rows 5-21, 17 items):
# Vapor + Lumber (sleepers/joists/rim/blocking) + Hardware + Insulation + Subfloor + Adhesives + Sealants + Thresholds + 4 NEW
RAISED_FLOOR_PRICES = [
    (5, 0.10, "Husky / generic 6-mil poly", "", "medium", "Confirmed", note_with_date("$0.10/sf. 10x100 roll ~$80-100")),
    (6, 0.95, "2x4 PT SPF/SYP", "", "medium", "Confirmed", note_with_date("PT для bearing на slab. 8ft ~$7-8")),
    (7, 1.85, "2x10 #2 SPF/HF 16 ft",
     "https://www.homedepot.com/p/2-in-x-10-in-x-16-ft-2-Premium-Grade-SPF-Dimensional-Lumber-201673/100016326",
     "high", "Confirmed", note_with_date("$1.85/LF. SPF 2x10 16ft ~$30/piece (Doug Fir ~$36.63/piece ref).")),
    (8, 2.20, "2x10 PT for perimeter", "", "medium", "Confirmed", note_with_date("PT для rim on concrete contact")),
    (9, 1.85, "2x10 SPF blocking", "", "medium", "Confirmed", note_with_date("Same as joists")),
    (10, 4.00, "Simpson LUS210", "", "medium", "Confirmed", note_with_date("Joist hanger for 2x10. $3.50-4.50 each")),
    (11, 1.50, "Tapcon 1/2\" wedge", "", "medium", "Confirmed", note_with_date("Concrete anchor PT to slab")),
    (12, 1.61, "Owens Corning R-30 EcoTouch 16\" OC unfaced",
     "https://www.homedepot.com/p/Owens-Corning-R-30-EcoTouch-PINK-Unfaced-Fiberglass-Insulation-Batt-16-in-x-48-in-BU74/300147480",
     "high", "Confirmed", note_with_date("$1.61/sf. R-30 unfaced 16x48. CA Title 24 compliant")),
    (13, 1.35, "CDX 3/4\" T&G plywood",
     "https://www.homedepot.com/p/3-4-in-x-4-ft-x-8-ft-CDX-Pine-Plywood-1060/202088096",
     "high", "Confirmed", note_with_date("$1.35/sf. 4x8 = $43/sheet (LENCO ref $44.95)")),
    (14, 10.00, "Loctite PL Premium 10oz", "", "medium", "Confirmed", note_with_date("Subfloor adhesive")),
    (15, 6.00, "GRK / generic 3\" #8", "", "medium", "Confirmed", note_with_date("Subfloor deck screws per lb. 5 lb box ~$30")),
    (16, 6.00, "Great Stuff window/door", "", "medium", "Confirmed", note_with_date("Foam sealant")),
    (17, 25.00, "Generic aluminum threshold 36\"", "", "medium", "TBD", note_with_date("Entry, French door, bathroom, exterior step")),
    (18, 12.00, "Owens Corning / TVM sill seal 1x50", "", "medium", "Confirmed", note_with_date("ДОДАНО: foam gasket між PT і slab")),
    (19, 35.00, "Mapei Planipatch 25 lb", "", "medium", "TBD", note_with_date("ДОДАНО: $35/bag, ~25 sf @ 1/8\". Партіальне выравнивание")),
    (20, 45.00, "DryLok concrete sealer 5 gal", "", "medium", "TBD", note_with_date("ДОДАНО: $45/gal, baseline qty=0 — only if §33.23 moisture issue")),
    (21, 12.00, "OSI ProSeries acoustical 28oz", "", "medium", "TBD", note_with_date("ДОДАНО: optional sound dampening")),
    # --- V3 reviewer additions ---
    (22, 15.00, "Stego seam tape / generic", "", "medium", "Confirmed", note_with_date("ДОДАНО: для overlap vapor barrier joints")),
    (23, 9.00, "Owens Corning insul support wires", "", "medium", "Confirmed", note_with_date("ДОДАНО: тримає R-30 batt між joists")),
    (24, 10.00, "Generic composite shims pack", "", "medium", "Confirmed", note_with_date("ДОДАНО: для leveling sleepers/joists")),
    (25, 25.00, "Simpson SDS / Ledgerlok box", "", "medium", "Confirmed", note_with_date("ДОДАНО: structural screws для rim-joist connections")),
    (26, 22.00, "Generic 12x12 PVC access panel", "", "medium", "TBD", note_with_date("ДОДАНО: MEP access під raised floor")),
]


# === 04_Framing — Phase 3 V2 (reviewer fixes + 6 new items) ===
# Структура (rows 5-26, 22 items): Studs + Plates + Headers + Blocking + Backing + Hardware + Anchors + Fasteners + Sheathing + Adhesives + Sealants + Fire + Layout
FRAMING_PRICES = [
    # --- STUDS ---
    (5, 4.50, "2x4 SPF #2 8ft", "", "medium", "Confirmed", note_with_date("$4.50/each Bay Area premium. Menards ref $3.64 + ~25% CA")),
    (6, 7.00, "2x6 SPF #2 8ft", "", "medium", "TBD", note_with_date("Wet wall studs — final per bathroom layout")),
    (7, 4.50, "2x4 SPF #2 8ft (jack/king/cripple)", "", "medium", "Confirmed", note_with_date("Same as standard 2x4, allowance для openings")),
    # --- PLATES ---
    (8, 0.95, "2x4 PT", "", "medium", "Confirmed", note_with_date("PT bottom plate per LF")),
    (9, 0.56, "2x4 SPF #2", "", "medium", "Confirmed", note_with_date("$0.56/LF ($4.50/8ft). Regular plate stock")),
    # --- HEADERS ---
    (10, 0.88, "2x6 SPF #2", "", "medium", "Confirmed", note_with_date("$0.88/LF ($7/8ft). Standard window/door headers")),
    (11, 6.00, "LVL 1.75x7.25 engineered",
     "https://www.homedepot.com/b/Lumber-Composites-Engineered-Wood-Products/LVL/N-5yc1vZcb73Z1z0xv6e",
     "medium", "TBD", note_with_date("LVL header для French door + folding 8ft. Final size per engineer")),
    # --- BLOCKING & BACKING ---
    (12, 0.56, "2x4 SPF #2 blocking", "", "medium", "Confirmed", note_with_date("Same as plates stock")),
    (13, 0.56, "2x4 SPF #2 nailer", "", "medium", "Confirmed", note_with_date("Drywall backing per LF")),
    # --- HARDWARE ---
    (14, 3.00, "Simpson mixed pack", "", "medium", "TBD", note_with_date("L, T, A35 mixed connectors")),
    (15, 30.00, "Simpson HDU2-SDS2.5", "", "medium", "TBD", note_with_date("Hold-down. $25-35 each. Allowance, TBD per engineer")),
    (16, 0.50, "Generic 16-gauge stud guard", "", "medium", "TBD", note_with_date("Nail protection plate for trades")),
    (17, 10.00, "Generic composite/wood shims pack", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): для wall plumbing і opening setting")),
    # --- ANCHORS ---
    (18, 1.00, "Simpson / generic 1/2 × 10 galv", "", "medium", "TBD", note_with_date("Anchor bolt — verify use for sill/perimeter only")),
    # --- FASTENERS ---
    (19, 5.00, "Generic 16d collated framing nails", "", "medium", "Confirmed", note_with_date("$5/lb pneumatic collated")),
    (20, 30.00, "Simpson SDS / Strong-Drive box", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): structural screws для LVL/header")),
    # --- SHEATHING ---
    (21, 30.00, "3/4\" OSB 4x8 sheet", "", "medium", "TBD", note_with_date("ДОДАНО (review): garage opening close-up + shear allowance")),
    # --- ADHESIVES & SEALANTS ---
    (22, 10.00, "Loctite PL Premium 10oz", "", "medium", "Confirmed", note_with_date("Plate-to-floor adhesive")),
    (23, 10.00, "OSI ProSeries acoustical 28oz", "", "medium", "TBD", note_with_date("Sound dampening для bedroom/bathroom partitions")),
    # --- FIRE PROTECTION ---
    (24, 30.00, "DAP fire-block foam + lumber", "", "medium", "TBD", note_with_date("Per code requirements")),
    (25, 12.00, "3M fire-rated caulk / Hilti FS-One", "", "medium", "TBD", note_with_date("ДОДАНО (review): окремо від foam blocking")),
    # --- LAYOUT SUPPLIES ---
    (26, 25.00, "Generic chalk line + pencils + plumb bob set", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): layout aids")),
    # --- BRACING ---
    # No new row for bracing — temporary bracing lumber in row 27
]


# Add row 27 for Temporary Bracing — but I just added 22 entries, row 26 is last
# Let me append it as row 27
FRAMING_PRICES.append(
    (27, 0.56, "2x4 SPF temp bracing (reusable)", "", "medium", "TBD", note_with_date("ДОДАНО (review): temporary bracing, можна reuse"))
)


# === 05_Doors_Windows — Phase 4 ===
# Структура: 16 existing + 8 new = 24 items, rows 5-28
DOORS_WINDOWS_PRICES = [
    # --- WINDOWS (3) ---
    (5, 180.00, "JELD-WEN V-4500 24×36 white",
     "https://www.homedepot.com/b/Doors-Windows-Windows-Single-Hung-Windows/Energy-Star/N-5yc1vZas73Z1z1wgir",
     "medium", "TBD", note_with_date("Bathroom window, smaller size ~$150-220")),
    (6, 300.00, "JELD-WEN V-4500 36×48 white",
     "https://www.homedepot.com/p/JELD-WEN-36-in-x-48-in-V-4500-Series-White-Single-Hung-Vinyl-Window-with-Fiberglass-Mesh-Screen-THDJW143900466/304848395",
     "medium", "TBD", note_with_date("Kitchen window, standard size. HD retail $250-350")),
    (7, 400.00, "JELD-WEN V-4500 36×60 white", "", "medium", "TBD", note_with_date("Living window, larger size. Premium V-4500 dealer $675 ref")),
    # --- EXTERIOR DOORS (2) ---
    (8, 400.00, "Masonite Premium 6-panel 36×80",
     "https://www.homedepot.com/p/Masonite-36-in-x-80-in-Premium-6-Panel-Right-Hand-Inswing-Primed-Steel-Prehung-Front-Exterior-Door-with-Brickmold-27105/100053184",
     "medium", "TBD", note_with_date("Prehung steel insulated з brickmold. $300-500 range")),
    (9, 1500.00, "Masonite/ThermaTru fiberglass 60×80 patio",
     "https://www.homedepot.com/b/Doors-Windows-Exterior-Doors-Patio-Doors/Masonite/60-x-80/French-Patio-Door/N-5yc1vZarqnZyqZ1z0rbkpZ1z0v1y3",
     "medium", "Confirmed", note_with_date("Special Order. Fiberglass French. $1200-1800 range")),
    # --- INTERIOR DOORS (2) ---
    (10, 110.00, "Masonite hollow core 30×80", "", "medium", "TBD", note_with_date("Prehung з privacy lock")),
    (11, 80.00, "Masonite louvered prehung 24×80", "", "medium", "TBD", note_with_date("Або bifold louvered")),
    # --- SPECIALTY SYSTEMS (2) ---
    (12, 250.00, "Home Depot bypass system 6ft", "", "medium", "TBD", note_with_date("2-panel bypass sliding closet system")),
    (13, 3000.00, "Mid-tier folding partition system 8ft",
     "https://homeguide.com/costs/folding-accordion-patio-doors-cost",
     "low", "Confirmed", note_with_date("Special Order INTERIOR partition. Variability $1500 (basic accordion) - $6000 (LaCantina-style interior). Mid $3000 для quality glass/wood folding")),
    # --- WEATHERPROOFING ---
    (14, 25.00, "Tyvek FlexWrap / 3M flashing tape", "", "medium", "Confirmed", note_with_date("Per roll")),
    (15, 30.00, "Generic vinyl sill pan", "", "medium", "Confirmed", note_with_date("Per window")),
    (16, 8.00, "Sika / OSI polyurethane sealant", "", "medium", "Confirmed", note_with_date("Caulk tube")),
    (17, 7.00, "Great Stuff window/door foam", "", "medium", "Confirmed", note_with_date("Low-expansion cans")),
    # --- TRIM ---
    (18, 0.80, "MDF / pine interior casing", "", "medium", "Confirmed", note_with_date("Per LF")),
    (19, 1.20, "Pine/PVC exterior trim", "", "medium", "Confirmed", note_with_date("Per LF")),
    # --- HARDWARE — REORGANIZED per reviewer V3 ---
    (20, 200.00, "Schlage/Kwikset Entry lockset+deadbolt", "", "medium", "Confirmed", note_with_date("Entry ONLY (French door has separate multi-point in row 24)")),
    # --- ДОДАНО (Pre-Takeoff Check) ---
    (21, 40.00, "Generic vinyl window screen", "", "medium", "Confirmed", note_with_date("Per window. May be included з window selection")),
    (22, 30.00, "Pine/MDF window sill stool", "", "medium", "Confirmed", note_with_date("Per window")),
    (23, 1.50, "Pine brick mold exterior casing", "", "medium", "Confirmed", note_with_date("Per LF, surrounds exterior openings")),
    (24, 300.00, "Multi-point lock French door set", "", "medium", "TBD", note_with_date("3-point lock + handles для French door")),
    (25, 25.00, "Generic kit weatherstripping", "", "medium", "Confirmed", note_with_date("Per exterior door")),
    (26, 20.00, "Generic door sweep / threshold seal", "", "medium", "Confirmed", note_with_date("Per exterior door")),
    (27, 50.00, "Generic threshold extender", "", "medium", "TBD", note_with_date("French door через raised floor 10\" rise")),
    (28, 80.00, "Generic astragal strip", "", "medium", "TBD", note_with_date("Якщо single-leaf French door")),
    # --- ДОДАНО per reviewer V3 ---
    (29, 30.00, "Schlage/Kwikset privacy or passage handle", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): bathroom privacy + utility passage")),
    (30, 25.00, "Generic 3-pack door hinges", "", "medium", "TBD", note_with_date("ДОДАНО (review): allowance")),
    (31, 10.00, "Generic composite/wood shims pack", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): door/window setting")),
    (32, 8.00, "Closed-cell backer rod 1/2\"", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): для exterior sealant gaps")),
    (33, 15.00, "Generic aluminum drip cap / Z-flashing", "", "medium", "TBD", note_with_date("ДОДАНО (review): над exterior openings")),
    (34, 5.00, "DAP Alex Plus paintable caulk", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): casing/trim joints")),
]


# === 06_Plumbing_Roughin — Phase 5 ===
# Структура (rows 5-31, 27 items): supply + drain + vent + fittings + misc + 10 new
PLUMBING_RI_PRICES = [
    # --- SUPPLY ---
    (5, 0.35, "SharkBite/Apollo PEX-B 1/2\" 100ft",
     "https://www.homedepot.com/p/SharkBite-1-2-in-x-100-ft-Coil-Red-PEX-B-Pipe-U860R100/202033011",
     "high", "Confirmed", note_with_date("$0.35/LF ($35/100ft coil)")),
    (6, 0.70, "SharkBite PEX-B 3/4\" 100ft", "", "medium", "Confirmed", note_with_date("$0.70/LF ($70/100ft coil)")),
    (7, 25.00, "Apollo/SharkBite PEX fittings pack", "", "medium", "Confirmed", note_with_date("Assortment: elbows, tees, couplings, crimp rings")),
    (8, 7.00, "Apollo PEX-B 1/2\" angle stop quarter-turn",
     "https://www.homedepot.com/p/Apollo-1-2-in-Chrome-Plated-Brass-PEX-B-Barb-x-1-4-in-Compression-Quarter-Turn-Angle-Stop-Valve-APXVA1214C/305455212",
     "high", "Confirmed", note_with_date("Per valve, brass chrome-plated")),
    # --- DRAIN ---
    (9, 0.85, "Charlotte/IPEX ABS DWV 1.5\"", "", "medium", "Confirmed", note_with_date("Per LF, DWV grade")),
    (10, 1.10, "Charlotte/IPEX ABS DWV 2\"", "", "medium", "Confirmed", note_with_date("Per LF")),
    (11, 1.80, "Charlotte/IPEX ABS DWV 3\"", "", "medium", "Confirmed", note_with_date("Per LF, for toilet drain")),
    (12, 3.00, "IPEX ABS DWV 4\" 10ft",
     "https://www.homedepot.com/p/IPEX-4-in-x-10-ft-ABS-DWV-Cell-Core-Pipe-179687/309282462",
     "medium", "Confirmed", note_with_date("$3/LF ($30/10ft piece). Cell core ABS")),
    # --- VENT ---
    (13, 0.95, "Charlotte ABS DWV 2\" vent", "", "medium", "Confirmed", note_with_date("Per LF")),
    # --- FITTINGS ---
    (14, 8.00, "Oatey/generic P-trap kit", "", "medium", "Confirmed", note_with_date("1.5\" or 2\" P-trap")),
    (15, 20.00, "Charlotte 4\" ABS cleanout", "", "medium", "Confirmed", note_with_date("Body + threaded plug")),
    (16, 15.00, "Sioux Chief strut/strap pack", "", "medium", "Confirmed", note_with_date("Per pack — pipe hangers/clamps")),
    (17, 12.00, "Oatey ABS primer + cement combo", "", "medium", "Confirmed", note_with_date("Per can")),
    # --- MISC ---
    (18, 40.00, "Sioux Chief / LSP washer outlet box", "", "medium", "Confirmed", note_with_date("Recessed hot/cold + drain box")),
    (19, 12.00, "Oatey ABS toilet flange w/ ring", "", "medium", "Confirmed", note_with_date("Standard 3\" or 4\"")),
    (20, 50.00, "WH connection kit (flex lines + dielectric)", "", "medium", "TBD", note_with_date("Coordinate Phase 15 WH")),
    (21, 0.50, "Armaflex foam 1/2\"-3/4\"", "", "medium", "Confirmed", note_with_date("Hot line insulation per LF")),
    # --- ДОДАНО (Pre-Takeoff Check) ---
    (22, 50.00, "ABS DWV fittings assortment pack",
     "", "medium", "Confirmed", note_with_date("ДОДАНО: elbows, tees, sanitary tees, wyes, reducers. $50/pack")),
    (23, 15.00, "2\" PVC/ABS washer standpipe", "", "medium", "Confirmed", note_with_date("ДОДАНО (§20.3)")),
    (24, 25.00, "Generic dishwasher install kit", "", "medium", "Confirmed", note_with_date("ДОДАНО: tee + elbow + supply line")),
    (25, 20.00, "Oatey No-Calk pipe boot 2\"", "", "medium", "Confirmed", note_with_date("ДОДАНО: vent stack roof penetration")),
    (26, 25.00, "Studor Mini AAV", "", "medium", "TBD", note_with_date("ДОДАНО: alternative venting allowance")),
    (27, 30.00, "Generic refrigerator water line kit", "", "medium", "TBD", note_with_date("ДОДАНО: 3/8 PEX + saddle + icemaker — optional")),
    (28, 12.00, "Sioux Chief mini-rester hammer arrestor", "", "medium", "Confirmed", note_with_date("ДОДАНО: washer + dishwasher")),
    (29, 15.00, "Rectorseal + Teflon tape combo", "", "medium", "Confirmed", note_with_date("ДОДАНО: threaded connections")),
    (30, 25.00, "Generic PEX-to-NPT brass adapters pack", "", "medium", "Confirmed", note_with_date("ДОДАНО: sink supply, WH connections")),
    (31, 25.00, "Generic stub-outs + escutcheons + test caps set", "", "medium", "Confirmed", note_with_date("ДОДАНО: finish + pressure test plugs")),
    # --- ДОДАНО per reviewer V2 ---
    (32, 70.00, "Delta MultiChoice / Moen Posi-Temp valve body", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): shower rough-in valve. Trim → Phase 13")),
    (33, 30.00, "Oatey 2\" shower drain assembly", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): body + strainer")),
    (34, 6.00, "Sioux Chief PEX drop-ear elbow", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): per pc")),
    (35, 20.00, "Generic copper crimp rings pack", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): extra crimps")),
    (36, 30.00, "Generic 30x32 washer drain pan", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): leak protection")),
    (37, 30.00, "Generic dishwasher branch + air gap", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): CA UPC air gap mandatory")),
    (38, 0.00, "—", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): cross-ref Phase 15. Не дублювати — qty 0")),
]


# === 08_HVAC_MiniSplit — Phase 7 ===
# IMPORTANT: Mini-split sold as kit (outdoor + 2 indoor + line sets).
# r5 = full kit price. r6 i r7 = qty 0 (included в kit).
HVAC_PRICES = [
    # --- OUTDOOR + KIT ---
    (5, 3150.00, "MRCOOL DIY 5th Gen 18k BTU 9k+9k kit (25ft lines)",
     "https://www.homedepot.com/p/MRCOOL-DIY-5th-Gen-18-000-BTU-1-5-Ton-2-Zone-23-9-SEER2-Ductless-Mini-Split-AC-and-Heat-Pump-with-9K-9K-and-25-25ft-Lines-DIYM218HPW03D06/337863265",
     "high", "Confirmed", note_with_date("$3,073 kit з 16ft lines + premium для 25ft = ~$3,150. Includes 2 indoor + line sets + controllers. Alt: Pioneer/Senville $2,500-2,800")),
    # --- INDOOR & LINE SETS — included в kit ---
    (6, 0.00, "Cross-ref kit r5 — qty 0", "", "high", "Confirmed", note_with_date("Included в MRCOOL DIY kit. Standalone $400/unit if separately purchased")),
    (7, 0.00, "Cross-ref kit r5 — qty 0", "", "high", "Confirmed", note_with_date("Included в kit. Standalone $100/set if separately")),
    # --- CONDENSATE ---
    (8, 30.00, "Generic 3/4 PVC + fittings", "", "medium", "Confirmed", note_with_date("Per set")),
    # --- PENETRATIONS ---
    (9, 30.00, "Generic 3-inch wall sleeve", "", "medium", "Confirmed", note_with_date("Per unit")),
    # --- MOUNTING ---
    (10, 35.00, "Quick-Sling wall bracket allowance", "", "medium", "Optional", note_with_date("Optional alternative — baseline = concrete pad r14")),
    (11, 70.00, "Mitsubishi/Pioneer line hide kit", "", "medium", "TBD", note_with_date("Decorative cover, 7-8 ft length")),
    (12, 40.00, "Generic lag bolts + brackets pack", "", "medium", "Confirmed", note_with_date("Mounting hardware")),
    (13, 0.00, "Included з indoor units", "", "high", "Confirmed", note_with_date("Standard remote included")),
    # --- ДОДАНО (Pre-Takeoff Check) ---
    (14, 35.00, "Diversitech precast 36x36 pad", "", "medium", "Confirmed", note_with_date("ДОДАНО: stable base + CA seismic")),
    (15, 20.00, "Diversitech rubber/cork pads", "", "medium", "Confirmed", note_with_date("ДОДАНО: noise/vibration")),
    (16, 15.00, "Generic 3/4 PVC condensate trap", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (17, 100.00, "Aspen Mini Aqua condensate pump", "", "medium", "Optional", note_with_date("ДОДАНО: optional, qty 0 baseline")),
    (18, 30.00, "Generic mesh filters set 4-pack", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (19, 50.00, "Cielo Breez Plus / generic Wi-Fi", "", "medium", "Optional", note_with_date("ДОДАНО: smart home")),
    (20, 30.00, "R-410A per lb", "", "medium", "Optional", note_with_date("ДОДАНО: qty 0 baseline (pre-charge sufficient)")),
    (21, 15.00, "Generic flare nuts/fittings pack", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (22, 10.00, "Armaflex insulation tape", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (23, 20.00, "Generic drip pan + leak sensor", "", "medium", "Optional", note_with_date("Optional. Drip pan only if compatible з indoor unit")),
    # --- ДОДАНО per reviewer V3 ---
    (24, 8.00, "GE Silicone II / DAP penetration sealant", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): per tube, exterior waterproofing")),
    (25, 12.00, "Generic line-set wall cap / penetration cover", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): per pc")),
    (26, 15.00, "Generic UV-rated zip ties + line clamps pack", "", "medium", "Confirmed", note_with_date("ДОДАНО (review)")),
    (27, 15.00, "Generic condensate termination fittings", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): exterior drain end")),
    (28, 0.00, "Cross-ref Phase 6 r19 — qty 0", "", "high", "Confirmed", note_with_date("ДОДАНО (review): electrical whip/disconnect у Phase 6 — не дублювати")),
]


# === 09_Insulation — Phase 8 ===
# Структура (rows 5-17, 13 items): Wall R-15 + Ceiling R-30 + Floor (cross-ref) + Sound R-13 + air sealing + 5 нових
INSULATION_PRICES = [
    (5, 1.00, "Owens Corning R-15 Kraft-Faced 16\" OC",
     "https://www.homedepot.com/p/Owens-Corning-R-15-Kraft-Faced-Fiberglass-Insulation-Batt-16-in-x-96-in-F95/313909974",
     "high", "Confirmed", note_with_date("$1.00/sf. R-15 kraft-faced range $0.69-1.10")),
    (6, 1.61, "Owens Corning R-30 Unfaced 16\" OC",
     "https://www.homedepot.com/p/Owens-Corning-R-30-EcoTouch-PINK-Unfaced-Fiberglass-Insulation-Batt-16-in-x-48-in-BU74/300147480",
     "high", "Confirmed", note_with_date("$1.61/sf. R-30 same Phase 2 pricing reference")),
    (7, 0.00, "Cross-ref Phase 2 row 12 — qty 0", "", "high", "Confirmed", note_with_date("DUPLICATE — already в Phase 2 Raised Floor")),
    (8, 1.10, "Owens Corning R-13 Kraft-Faced",
     "https://www.homedepot.com/p/Owens-Corning-R-13-Kraft-Faced-Fiberglass-Insulation-Batt-16-in-x-96-in-10-Bags-C86/202746452",
     "medium", "Confirmed", note_with_date("$1.10/sf fiberglass R-13. Alt mineral wool $1.85/sf r15")),
    (9, 7.00, "Great Stuff Pro window/door foam", "", "medium", "Confirmed", note_with_date("Per can")),
    (10, 10.00, "DAP Fire-block foam tube", "", "medium", "Confirmed", note_with_date("Per tube")),
    (11, 25.00, "3M All Weather flashing tape", "", "medium", "Confirmed", note_with_date("Per roll")),
    (12, 1.50, "ADO Products attic rafter vent baffle", "", "medium", "TBD", note_with_date("Per pc")),
    # --- ДОДАНО (Pre-Takeoff Check) ---
    (13, 9.00, "Owens Corning insulation wire supports", "", "medium", "Confirmed", note_with_date("ДОДАНО: per pack")),
    (14, 40.00, "Stego/generic 6-mil poly allowance", "", "medium", "Optional", note_with_date("ДОДАНО: kraft-faced usually sufficient")),
    (15, 1.85, "ROCKWOOL Safe'n'Sound R-13",
     "https://www.homedepot.com/p/ROCKWOOL-Safe-n-Sound-3-in-x-15-1-4-in-x-47-in-Soundproofing-and-Fire-Resistant-Stone-Wool-Insulation-Batt-59-7-sq-ft-RXSS31525/202531875",
     "high", "Optional", note_with_date("$1.85/sf. Premium mineral wool alternative to r8. Choose ONE not both")),
    (16, 10.00, "Generic cap strips / batt straps pack", "", "medium", "TBD", note_with_date("ДОДАНО")),
    (17, 300.00, "Tiger Foam closed-cell kit (Optional)", "", "low", "Optional", note_with_date("ДОДАНО: closed-cell spray foam allowance")),
    # --- ДОДАНО per reviewer V3 ---
    (18, 5.00, "DAP Alex Plus general caulk", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): top/bottom plates, corners")),
    (19, 20.00, "Generic T50 staples + insulation fasteners set", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): kraft-faced batt stapling")),
    (20, 25.00, "Generic PPE pack (gloves + mask + goggles)", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): fiberglass safety")),
]


# === 10_Drywall — Phase 9 ===
# Структура (rows 5-19, 15 items): board (regular + moisture + cement) + fasteners + finishing + 7 нових
DRYWALL_PRICES = [
    (5, 13.00, "USG UltraLight 1/2\" 4x8 sheet",
     "https://www.homedepot.com/p/USG-Sheetrock-Brand-1-2-in-x-4-ft-x-8-ft-UltraLight-Drywall-14113411708/202530243",
     "high", "Confirmed", note_with_date("$11-15/sheet. ToughRock alt cheaper")),
    (6, 20.00, "USG Mold Tough 1/2\" 4x8 greenboard",
     "https://www.homedepot.com/p/USG-Sheetrock-Brand-1-2-in-x-4-ft-x-8-ft-UltraLight-Mold-Tough-Drywall-14302111708/203768542",
     "high", "Confirmed", note_with_date("$22.98 single / $19.53 bulk 34+. Mid $20")),
    (7, 20.00, "HardieBacker 1/2\" 3x5 cement board",
     "https://www.homedepot.com/p/James-Hardie-HardieBacker-1-2-in-x-3-ft-x-5-ft-Cement-Backerboard-220023/100170507",
     "medium", "Confirmed", note_with_date("Range $13-25 depending on size 3x5 vs 4x8")),
    (8, 7.00, "Generic 1¼ coarse drywall screws", "", "medium", "Confirmed", note_with_date("Per lb. 5 lb box ~$35")),
    (9, 17.00, "USG Sheetrock All-Purpose 5 gal", "", "medium", "Confirmed", note_with_date("Per bucket")),
    (10, 5.00, "Generic paper joint tape 250 ft", "", "medium", "Confirmed", note_with_date("Per roll")),
    (11, 0.65, "Generic metal corner bead 8 ft", "", "medium", "Confirmed", note_with_date("Per LF")),
    (12, 25.00, "Kilz PVA Sealer primer 1 gal", "", "medium", "Confirmed", note_with_date("Per gal")),
    # --- ДОДАНО (Pre-Takeoff Check) ---
    (13, 30.00, "Homax knockdown texture spray", "", "medium", "TBD", note_with_date("ДОДАНО (§24.4): per can. CA default knockdown")),
    (14, 22.00, "Generic PVC access panel 12x12", "", "medium", "Confirmed", note_with_date("ДОДАНО (§24.4): plumbing/electric access")),
    (15, 5.00, "Generic alkali-resistant mesh tape", "", "medium", "Confirmed", note_with_date("ДОДАНО: cement board joints")),
    (16, 50.00, "Custom Building Products RedGard 1 gal",
     "https://www.homedepot.com/b/Flooring-Flooring-Supplies/Custom-Building-Products/Waterproofing-Membrane/N-5yc1vZcabsZ1z10jx7Z2g1",
     "medium", "Confirmed", note_with_date("ДОДАНО: waterproofing за cement board у shower")),
    (17, 20.00, "Generic sanding sponges + screens + paper set", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (18, 25.00, "USG Durabond 90 setting-type 25 lb", "", "medium", "Optional", note_with_date("ДОДАНО: 20-min set для repair/shower")),
    (19, 10.00, "Loctite PL Premium 10oz drywall adhesive", "", "medium", "Optional", note_with_date("ДОДАНО: stud attachment optional")),
    # --- ДОДАНО per reviewer V3 ---
    (20, 18.00, "5/8\" Type X drywall 4x8 sheet", "", "medium", "Optional", note_with_date("ДОДАНО (review): fire-rated upgrade. Qty 0 baseline")),
    (21, 10.00, "No-Coat flexible corner tape 100 ft", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): vaulted ceiling")),
    (22, 30.00, "Generic 10x100 plastic sheeting 3-mil", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): dust protection")),
    (23, 1.50, "Trim-Tex control joint vinyl", "", "medium", "Optional", note_with_date("ДОДАНО (review): per LF, qty 0 baseline")),
    (24, 5.00, "Generic 6x6 drywall patch", "", "medium", "Optional", note_with_date("ДОДАНО (review): qty 0 baseline")),
]


# === 11_Paint — Phase 10 ===
# IMPORTANT: PVA drywall primer counted в Phase 9 r12 — НЕ дублювати
PAINT_PRICES = [
    (5, 35.00, "Behr Premium Plus Ultra eggshell 1 gal", "", "medium", "Confirmed", note_with_date("Range $30-45/gal. 5gal bucket cheaper per gal")),
    (6, 30.00, "Behr Premium Plus Ultra ceiling flat 1 gal", "", "medium", "Confirmed", note_with_date("Flat ceiling-specific")),
    (7, 45.00, "Behr Premium Mildew-resistant semi-gloss 1 gal", "", "medium", "Confirmed", note_with_date("Bathroom moisture")),
    (8, 42.00, "Behr Premium semi-gloss acrylic 1 gal", "", "medium", "Confirmed", note_with_date("Trim/door paint")),
    (9, 5.00, "DAP Alex Plus painter's caulk", "", "medium", "Confirmed", note_with_date("Per tube")),
    (10, 12.00, "3M ScotchBlue painter's tape 1.88\"", "", "medium", "Confirmed", note_with_date("Per roll")),
    (11, 30.00, "Generic canvas drop cloths + plastic set", "", "medium", "Confirmed", note_with_date("Mixed canvas + plastic sheeting")),
    (12, 20.00, "Generic sandpaper + DAP wood filler set", "", "medium", "Confirmed", note_with_date("Assortment grit + filler")),
    # --- ДОДАНО (Pre-Takeoff Check) ---
    (13, 15.00, "Generic touch-up supplies kit", "", "medium", "Confirmed", note_with_date("ДОДАНО (per §27.2)")),
    (14, 20.00, "Wooster 9\" roller covers 6-pack 3/8 nap", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (15, 25.00, "Purdy paint brushes set 2\"+2.5\"+3\"", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (16, 15.00, "9\" paint tray + 4-pack disposable liners", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (17, 20.00, "Generic telescoping roller pole 4-8 ft", "", "medium", "Confirmed", note_with_date("ДОДАНО: vaulted ceiling")),
    (18, 8.00, "DAP DryDex spackling compound 1 lb", "", "medium", "Confirmed", note_with_date("ДОДАНО: малі отвори")),
    (19, 25.00, "Kilz Premium oil-based stain blocker 1 gal", "", "medium", "Confirmed", note_with_date("ДОДАНО: finish primer окремо від Phase 9 PVA")),
    (20, 15.00, "Mineral spirits 1 gal + Goof Off + rags", "", "medium", "Confirmed", note_with_date("ДОДАНО: cleanup combo")),
    (21, 15.00, "Generic putty knife + scraper set", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    # --- ДОДАНО per reviewer V2 ---
    (22, 7.00, "GE Silicone II 100% silicone caulk", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): wet joints bathroom/laundry")),
    (23, 15.00, "Trimaco pre-taped plastic masking film", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): cabinet/floor protection")),
    (24, 10.00, "Generic tack cloth + microfiber cloths", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): trim prep")),
]


# === 12_Flooring_Trim — Phase 11 ===
FLOORING_TRIM_PRICES = [
    (5, 3.00, "LifeProof 22 MIL Click-Lock Waterproof LVP",
     "https://www.homedepot.com/p/Lifeproof-Sterling-Oak-22-MIL-x-8-7-in-W-x-48-in-L-Click-Lock-Waterproof-Luxury-Vinyl-Plank-Flooring-20-1-sqft-case-I966106LP/309083456",
     "high", "Confirmed", note_with_date("$2.98-3.50/sf. LifeProof Sterling Oak / Trail / Fresh — Home Depot exclusive (Shaw)")),
    (6, 0.30, "Roberts Super Felt foam underlayment", "", "medium", "TBD", note_with_date("Per sf. Verify if LVP має attached underlayment first")),
    (7, 20.00, "Generic LVP T-molding/reducer/end-cap", "", "medium", "Confirmed", note_with_date("Per piece. Average mix")),
    (8, 0.85, "Pine MDF 5.25\" primed baseboard", "", "medium", "Confirmed", note_with_date("Per LF")),
    (9, 0.55, "Pine 3/4\" shoe molding primed", "", "medium", "TBD", note_with_date("Per LF")),
    # --- ДОДАНО (Pre-Takeoff Check) ---
    (10, 0.10, "Generic 6-mil poly vapor sheet", "", "medium", "Optional", note_with_date("ДОДАНО: only if LVP без attached vapor")),
    (11, 5.00, "Generic flooring spacers 1/4-3/8\" pack", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (12, 20.00, "Roberts LVP installation kit", "", "medium", "Confirmed", note_with_date("ДОДАНО: tapping block + pull bar + utility knife")),
    (13, 5.00, "Generic floor caulk tube", "", "medium", "Confirmed", note_with_date("ДОДАНО: perimeter sealing")),
    (14, 15.00, "Generic 18-gauge brad nails box", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (15, 20.00, "Door undercut allowance (saw rental or service)", "", "medium", "TBD", note_with_date("ДОДАНО")),
    (16, 25.00, "Mapei Planipatch 25 lb (Optional)", "", "medium", "Optional", note_with_date("ДОДАНО: small allowance. Cross-ref Phase 2 r19")),
    # --- ДОДАНО per reviewer V2 ---
    (17, 25.00, "Generic bathroom reducer / transition", "", "medium", "TBD", note_with_date("ДОДАНО (review): per piece")),
    (18, 40.00, "Generic LVP stair nosing / step nose", "", "medium", "TBD", note_with_date("ДОДАНО (review): через 10\" raised floor step")),
]


# === 13_Kitchen_Cabinets — Phase 12 ===
KITCHEN_CABINETS_PRICES = [
    (5, 50.00, "IKEA SEKTION base cabinet 24\" deep",
     "https://www.ikea.com/us/en/cat/base-cabinets-frame-height-23607/",
     "medium", "Confirmed", note_with_date("Per LF carcass ~$50/LF (3 cabinets × $100-130 / 8.5 LF)")),
    (6, 35.00, "IKEA SEKTION wall cabinet 15\" deep",
     "https://www.ikea.com/us/en/p/sektion-wall-cabinet-white-70265469/",
     "high", "Confirmed", note_with_date("Per LF ($78 SEKTION 30\" wide / 2.5 LF = $31/LF). Mid $35")),
    (7, 200.00, "IKEA SEKTION tall pantry 24\"x80\"", "", "medium", "TBD", note_with_date("Carcass пантри cabinet")),
    (8, 400.00, "IKEA METOD/AXSTAD doors set (mix)", "", "medium", "TBD", note_with_date("6-7 doors mix sizes")),
    (9, 240.00, "IKEA drawer fronts set", "", "medium", "TBD", note_with_date("6 drawers")),
    (10, 40.00, "IKEA UTRUSTA soft-close hinges pack", "", "medium", "TBD", note_with_date("")),
    (11, 40.00, "IKEA MAXIMERA soft-close drawer runners set", "", "medium", "TBD", note_with_date("")),
    (12, 25.00, "IKEA adjustable leveling legs set", "", "medium", "TBD", note_with_date("")),
    (13, 4.00, "IKEA FÖRBÄTTRA toe-kick per LF", "", "medium", "TBD", note_with_date("")),
    (14, 80.00, "IKEA FÖRBÄTTRA cover panels set", "", "medium", "TBD", note_with_date("")),
    (15, 50.00, "IKEA FÖRBÄTTRA filler panels set", "", "medium", "TBD", note_with_date("")),
    (16, 26.00, "IKEA KARLBY butcher block 9.5 LF",
     "https://www.ikea.com/us/en/cat/countertops-10653/",
     "medium", "TBD", note_with_date("Per LF. ~$250 for 9.5 LF. EKBACKEN laminate cheaper ~$14/LF")),
    (17, 4.00, "Generic ceramic subway 3x6 white",
     "", "medium", "TBD", note_with_date("Per sf. $4/sf retail")),
    (18, 175.00, "IKEA HILLESJÖN 25\" SS single bowl drop-in", "", "medium", "TBD", note_with_date("Or LANGUDDEN")),
    (19, 110.00, "IKEA ÄLMAREN pull-down faucet", "", "medium", "TBD", note_with_date("")),
    (20, 50.00, "IKEA modern bar pulls set 6-8 pcs", "", "medium", "TBD", note_with_date("")),
    (21, 80.00, "IKEA OMLOPP LED under-cabinet set", "", "medium", "TBD", note_with_date("Driver окремо в Phase 6 r38")),
    # --- ДОДАНО (Pre-Takeoff Check) ---
    (22, 20.00, "Oatey sink drain assembly + strainer kit", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (23, 15.00, "Generic drop-in sink clips або undermount brackets", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (24, 30.00, "Custom Building Products tile mastic + grout set", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (25, 30.00, "Generic bullnose backsplash trim set", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (26, 8.00, "GE Silicone II Kitchen 100% silicone", "", "medium", "Confirmed", note_with_date("ДОДАНО: per tube")),
    (27, 60.00, "IKEA RATIONELL/SEKTION extras (Optional)", "", "medium", "Optional", note_with_date("ДОДАНО: drawer inserts, extra shelves")),
    (28, 30.00, "IKEA cabinet feet/risers allowance (Optional)", "", "medium", "Optional", note_with_date("ДОДАНО")),
    # --- ДОДАНО per reviewer V2 ---
    (29, 12.00, "IKEA SEKTION suspension rail 39\"", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): per rail")),
    (30, 15.00, "Generic cabinet mounting screws + anchors", "", "medium", "Confirmed", note_with_date("ДОДАНО (review)")),
    (31, 20.00, "Generic countertop end caps + edge banding", "", "medium", "TBD", note_with_date("ДОДАНО (review)")),
    (32, 50.00, "IKEA dishwasher cover end panel", "", "medium", "TBD", note_with_date("ДОДАНО (review)")),
    (33, 15.00, "Generic dishwasher anti-tip mounting", "", "medium", "TBD", note_with_date("ДОДАНО (review)")),
    (34, 15.00, "Generic tile spacers + float + sponge kit", "", "medium", "Confirmed", note_with_date("ДОДАНО (review)")),
    (35, 20.00, "Boos butcher block oil + sealer 1 qt", "", "medium", "TBD", note_with_date("ДОДАНО (review): тільки для butcher block")),
    (36, 0.00, "Cross-ref Phase 5 r37 — qty 0", "", "high", "Confirmed", note_with_date("ДОДАНО: dishwasher air gap — already in Phase 5")),
]


# === 14_Bathroom_Fixtures — Phase 13 ===
BATHROOM_FIXTURES_PRICES = [
    (5, 275.00, "MAAX/Sterling prefab shower base 32×40", "", "medium", "Confirmed", note_with_date("Acrylic/fiberglass base only")),
    (6, 0.00, "DEPRECATED — заміняємо tile shower", "", "high", "Confirmed", note_with_date("Removed: заміняємо на tile shower walls (r38+)")),
    (7, 0.00, "Cross-ref Phase 5 r33", "", "high", "Confirmed", note_with_date("Shower drain — Phase 5")),
    (8, 0.00, "Cross-ref Phase 5 r32", "", "high", "Confirmed", note_with_date("Shower rough-in valve body — Phase 5")),
    (9, 80.00, "Delta MultiChoice trim kit", "", "medium", "TBD", note_with_date("Visible handle + escutcheon plate")),
    (10, 50.00, "Generic shower head fixed або handheld", "", "medium", "TBD", note_with_date("")),
    (11, 7.00, "GE Silicone II 100% silicone bath caulk", "", "medium", "Confirmed", note_with_date("Per tube")),
    (12, 200.00, "Glacier Bay 2-piece elongated chair-height", "", "medium", "Confirmed", note_with_date("Basic toilet $150-300")),
    (13, 8.00, "Generic braided supply line 12\"", "", "medium", "Confirmed", note_with_date("")),
    (14, 200.00, "Home Decorators Doveton 24\" vanity",
     "https://www.homedepot.com/p/Home-Decorators-Collection-Doveton-24-in-Single-Sink-Freestanding-White-Bath-Vanity-with-White-Engineered-Marble-Top-Assembled-Doveton-24W/324252515",
     "high", "Confirmed", note_with_date("$299-349 combo з top+sink (split here as r14+r15)")),
    (15, 150.00, "Engineered marble top + integrated sink", "", "medium", "Confirmed", note_with_date("Combined HD price з vanity")),
    (16, 100.00, "Glacier Bay/Moen single-handle vanity faucet", "", "medium", "Confirmed", note_with_date("")),
    (17, 8.00, "Generic 1.5\" P-trap kit", "", "medium", "Confirmed", note_with_date("Bathroom small allowance")),
    (18, 7.00, "Apollo PEX angle stop 1/2\"", "", "medium", "Confirmed", note_with_date("Per valve")),
    (19, 60.00, "Generic mirror 24x30 framed", "", "medium", "Confirmed", note_with_date("")),
    (20, 40.00, "Generic 24\" towel bar + 2 hooks set", "", "medium", "Confirmed", note_with_date("")),
    (21, 20.00, "Generic toilet paper holder", "", "medium", "Confirmed", note_with_date("")),
    (22, 100.00, "Broan 80 CFM exhaust fan", "", "medium", "Confirmed", note_with_date("Energy Star bath fan")),
    (23, 30.00, "Generic 4\" flex duct + wall cap", "", "medium", "Confirmed", note_with_date("")),
    # --- ДОДАНО (Pre-Takeoff Check) ---
    (24, 35.00, "Generic shower curtain rod + curtain", "", "medium", "Confirmed", note_with_date("ДОДАНО: budget enclosure")),
    (25, 7.00, "Oatey wax ring + sealing kit", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (26, 30.00, "Bemis/generic toilet seat", "", "medium", "Confirmed", note_with_date("ДОДАНО: NOT included з toilet")),
    (27, 5.00, "Generic toilet bolts + caps", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (28, 15.00, "Generic vanity wall mount anchors", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (29, 50.00, "Vanity backsplash tile (Optional)", "", "medium", "Optional", note_with_date("ДОДАНО")),
    (30, 15.00, "Generic robe hook", "", "medium", "Optional", note_with_date("ДОДАНО")),
    # --- ДОДАНО per reviewer V2 ---
    (31, 15.00, "Generic vanity pop-up drain assembly", "", "medium", "Confirmed", note_with_date("ДОДАНО (review)")),
    (32, 5.00, "Generic flex faucet supply line", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): per piece, mid 2")),
    (33, 15.00, "Generic shower arm + flange", "", "medium", "Confirmed", note_with_date("ДОДАНО (review)")),
    (34, 10.00, "Generic duct clamps + foil HVAC tape", "", "medium", "Confirmed", note_with_date("ДОДАНО (review)")),
    (35, 5.00, "Generic shower curtain hooks/rings", "", "medium", "Confirmed", note_with_date("ДОДАНО (review)")),
    (36, 80.00, "Generic medicine cabinet 16x20", "", "medium", "Optional", note_with_date("ДОДАНО (review)")),
    (37, 30.00, "Moen/generic grab bar 18-24\"", "", "medium", "Optional", note_with_date("ДОДАНО (review): blocking у Phase 3")),
    # --- ДОДАНО: tile shower walls ---
    (38, 3.00, "Ceramic subway 3×6 white", "", "medium", "Confirmed", note_with_date("ДОДАНО: $3/sf retail")),
    (39, 35.00, "Custom Building Products thinset 50 lb", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (40, 20.00, "Custom Building Products sanded grout 25 lb", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (41, 35.00, "Schluter trim / metal edges set", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    # --- ДОДАНО: tile floor ---
    (42, 4.00, "Porcelain floor tile 12×12", "", "medium", "Confirmed", note_with_date("ДОДАНО: $4/sf")),
    (43, 35.00, "Floor thinset 50 lb", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (44, 20.00, "Floor grout + spacers + sponge kit", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (45, 0.50, "HardieBacker 1/2\" cement board", "", "medium", "Confirmed", note_with_date("ДОДАНО: per sf для floor")),
    # --- ДОДАНО: waterproofing extra ---
    (46, 30.00, "RedGard quart (floor + corners)", "", "medium", "Confirmed", note_with_date("ДОДАНО: окремий quart")),
]


# === 15_Closet_System — Phase 14 ===
CLOSET_PRICES = [
    (5, 262.00, "IKEA BOAXEL 49\" wardrobe combination",
     "https://www.ikea.com/us/en/p/boaxel-wardrobe-combination-white-s19332373/",
     "high", "Confirmed", note_with_date("BOAXEL combo includes frame+shelves+rail")),
    (6, 10.00, "IKEA BOAXEL clothes rail 24\"-31\"", "", "medium", "TBD", note_with_date("Per pc")),
    (7, 15.00, "IKEA BOAXEL extra shelf", "", "medium", "TBD", note_with_date("Per pc")),
    (8, 20.00, "IKEA BOAXEL mesh basket drawer", "", "medium", "TBD", note_with_date("Per pc")),
    (9, 30.00, "IKEA BOAXEL shoe shelf set", "", "medium", "TBD", note_with_date("")),
    (10, 40.00, "IKEA wall-mounted mirror full-length", "", "medium", "TBD", note_with_date("Optional but Mid 1")),
    # --- ДОДАНО (Pre-Takeoff Check) ---
    (11, 20.00, "IKEA BOAXEL wall rail (additional)", "", "medium", "Confirmed", note_with_date("ДОДАНО: per rail")),
    (12, 15.00, "Generic drywall anchors + screws set", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (13, 20.00, "IKEA KOMPLEMENT drawer inserts (Optional)", "", "medium", "Optional", note_with_date("ДОДАНО")),
    (14, 15.00, "Closet accessories valet/belt rack (Optional)", "", "medium", "Optional", note_with_date("ДОДАНО")),
    # --- ДОДАНО per reviewer V2 ---
    (15, 15.00, "IKEA BOAXEL upright / vertical rail", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): per upright")),
    (16, 5.00, "IKEA BOAXEL shelf/rod brackets pack", "", "medium", "Confirmed", note_with_date("ДОДАНО (review)")),
    (17, 80.00, "IKEA BOAXEL second-wall add-on combo", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): для 5'6\" wall")),
    (18, 350.00, "IKEA PAX wardrobe upgrade alternative", "", "medium", "Optional", note_with_date("ДОДАНО (review): premium upgrade. Qty 0 baseline")),
    (19, 0.00, "Cross-ref Phase 4 r8 — qty 0", "", "high", "Confirmed", note_with_date("ДОДАНО: closet sliding door in Phase 4")),
    (20, 0.00, "Cross-ref Phase 6 r25 — qty 0", "", "high", "Confirmed", note_with_date("ДОДАНО: closet motion light in Phase 6")),
    (21, 20.00, "Generic plastic hangers set 50pc", "", "medium", "Optional", note_with_date("ДОДАНО (review)")),
]


# === 16_Water_Heater — Phase 15 ===
WATER_HEATER_PRICES = [
    (5, 429.00, "Rheem Performance 30 Gal Short 240V",
     "https://www.homedepot.com/p/Rheem-Performance-30-Gal-4500-Watt-Elements-Short-Electric-Water-Heater-with-6-Year-Tank-Warranty-and-240-Volt-XE30S06ST45U1/326590405",
     "high", "Confirmed", note_with_date("$379 promo / $429 regular. 4500W 6-yr warranty")),
    (6, 25.00, "Camco 24\" round WH pan", "", "medium", "Confirmed", note_with_date("")),
    (7, 5.00, "3/4\" PVC drain pipe per ft", "", "medium", "Confirmed", note_with_date("3/4\" consistent з spec")),
    (8, 0.00, "Cross-ref Phase 5 r20", "", "high", "Confirmed", note_with_date("Flex connectors у Phase 5 WH Connection Kit")),
    (9, 0.00, "Cross-ref Phase 5 r8", "", "high", "Confirmed", note_with_date("Shut-offs у Phase 5 (19 total)")),
    (10, 5.00, "Cu 3/4\" T&P discharge pipe per ft", "", "medium", "Confirmed", note_with_date("")),
    (11, 50.00, "Watts thermal expansion tank", "", "medium", "TBD", note_with_date("If req'd by code")),
    (12, 25.00, "Holdrite/generic seismic strap kit", "", "medium", "Confirmed", note_with_date("CA mandatory")),
    (13, 0.00, "Cross-ref Phase 4 r11", "", "high", "Confirmed", note_with_date("Utility closet door already у Phase 4 $80")),
    (14, 20.00, "Generic ventilation louver grille", "", "medium", "Confirmed", note_with_date("")),
    # --- ДОДАНО (Pre-Takeoff Check) ---
    (15, 15.00, "Generic pan drain termination + cap", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (16, 25.00, "Watts T&P relief valve replacement", "", "medium", "Optional", note_with_date("ДОДАНО: usually included з WH")),
    (17, 7.00, "Watts dielectric union 3/4\"", "", "medium", "Confirmed", note_with_date("ДОДАНО: per piece, qty 2")),
    # --- ДОДАНО per reviewer V2 ---
    (18, 10.00, "Generic lag screws + fasteners set", "", "medium", "Confirmed", note_with_date("ДОДАНО (review)")),
    (19, 12.00, "Generic T&P discharge termination fitting", "", "medium", "Confirmed", note_with_date("ДОДАНО (review)")),
    (20, 15.00, "Generic drain hose + service cap (Optional)", "", "medium", "Optional", note_with_date("ДОДАНО (review)")),
    (21, 0.00, "Cross-ref Phase 6 — qty 0", "", "high", "Confirmed", note_with_date("ДОДАНО (review): WH electrical у Phase 6")),
]


# === 17_Appliances — Phase 16 ===
APPLIANCES_PRICES = [
    (5, 500.00, "Frigidaire FFTR1022QS 24\" compact fridge", "", "medium", "Confirmed", note_with_date("$400-650 range. 10 cu ft compact")),
    (6, 650.00, "Frigidaire FFEH2422US 24\" electric range",
     "https://www.homedepot.com/p/Frigidaire-24-in-1-9-cu-ft-4-Burner-Element-Freestanding-Electric-Range-with-Manual-Clean-in-Stainless-Steel-FFEH2422US/309052067",
     "medium", "Confirmed", note_with_date("$500-800 range")),
    (7, 600.00, "Bosch 100 series 24\" compact dishwasher", "", "medium", "Confirmed", note_with_date("$450-900 range. Bosch quiet")),
    (8, 300.00, "GE JVM3160DFBB OTR microwave/hood", "", "medium", "Confirmed", note_with_date("$200-400 range")),
    (9, 115.00, "InSinkErator Badger 1/3 HP disposal", "", "medium", "Optional", note_with_date("Qty 0 baseline — Optional per ТЗ")),
    (10, 1000.00, "Bosch 500 WAT28402UC 24\" front-load", "", "medium", "Confirmed", note_with_date("Special Order. $850-1,200 range")),
    (11, 1400.00, "Bosch 500 WTG86402UC ventless heat pump", "", "medium", "Confirmed", note_with_date("Special Order. $1,200-1,700 range")),
    (12, 90.00, "Bosch stacking kit WTZ20410", "", "medium", "Confirmed", note_with_date("Brand-matched")),
    # --- ДОДАНО (Pre-Takeoff Check) ---
    (13, 50.00, "Generic 4\" round duct + exterior cap", "", "medium", "Confirmed", note_with_date("ДОДАНО: microwave/hood vent")),
    (14, 30.00, "Generic disposal install kit (Optional)", "", "medium", "Optional", note_with_date("ДОДАНО")),
    (15, 35.00, "Generic disposal air switch (Optional)", "", "medium", "Optional", note_with_date("ДОДАНО")),
    (16, 25.00, "Generic 14-50R range power cord 6 ft", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    (17, 25.00, "Generic 4-prong dryer cord", "", "medium", "Confirmed", note_with_date("ДОДАНО")),
    # --- ДОДАНО per reviewer V2 ---
    (18, 25.00, "Generic dishwasher install kit", "", "medium", "Confirmed", note_with_date("ДОДАНО (review): water line + cord + drain")),
    (19, 15.00, "Stainless braided washer hose pair", "", "medium", "Confirmed", note_with_date("ДОДАНО (review)")),
    (20, 15.00, "Generic dryer condensate drain kit", "", "medium", "TBD", note_with_date("ДОДАНО (review)")),
    (21, 20.00, "Generic OTR rectangular-to-round adapter + tape", "", "medium", "Confirmed", note_with_date("ДОДАНО (review)")),
]


UPDATES = {
    "18_Furniture_IKEA": FURNITURE_PRICES,
    "07_Electrical_Roughin": ELECTRICAL_PRICES,
    "02_Site_Exterior_Utilities": SITE_EXTERIOR_PRICES,
    "03_Raised_Floor_System": RAISED_FLOOR_PRICES,
    "04_Framing": FRAMING_PRICES,
    "05_Doors_Windows": DOORS_WINDOWS_PRICES,
    "06_Plumbing_Roughin": PLUMBING_RI_PRICES,
    "08_HVAC_MiniSplit": HVAC_PRICES,
    "09_Insulation": INSULATION_PRICES,
    "10_Drywall": DRYWALL_PRICES,
    "11_Paint": PAINT_PRICES,
    "12_Flooring_Trim": FLOORING_TRIM_PRICES,
    "13_Kitchen_Cabinets": KITCHEN_CABINETS_PRICES,
    "14_Bathroom_Fixtures": BATHROOM_FIXTURES_PRICES,
    "15_Closet_System": CLOSET_PRICES,
    "16_Water_Heater": WATER_HEATER_PRICES,
    "17_Appliances": APPLIANCES_PRICES,
}

# Корекції quantity (де комбо-ціна = "пакет = 1" або cross-ref kit)
QTY_OVERRIDES = {
    "18_Furniture_IKEA": {
        10: (1, 1, 1),  # Curtains combo як 1 пакет
    },
    "08_HVAC_MiniSplit": {
        6: (0, 0, 0),  # Indoor units — included in MRCOOL DIY kit
        7: (0, 0, 0),  # Line sets — included in kit
    },
}


def update_sheet(ws, rows, qty_overrides=None):
    n = 0
    for row, price, sku, url, conf, status, notes_add in rows:
        ws.cell(row=row, column=10, value=price)  # J: Unit Price
        ws.cell(row=row, column=10).number_format = "$#,##0.00"
        ws.cell(row=row, column=13, value=sku)  # M: SKU/Model
        ws.cell(row=row, column=14, value=url)  # N: URL
        if status is not None:
            ws.cell(row=row, column=15, value=status)  # O: Status
        ws.cell(row=row, column=16, value=conf)  # P: Confidence
        if notes_add is not None:
            existing = ws.cell(row=row, column=17).value or ""
            merged = f"{existing} | {notes_add}" if existing else notes_add
            ws.cell(row=row, column=17, value=merged)
        if qty_overrides and row in qty_overrides:
            ql, qm, qh = qty_overrides[row]
            ws.cell(row=row, column=7, value=ql)
            ws.cell(row=row, column=8, value=qm)
            ws.cell(row=row, column=9, value=qh)
        n += 1
    return n


def main():
    wb = load_workbook(XLSX)
    total_updated = 0
    for sheet_name, rows in UPDATES.items():
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  Лист {sheet_name} не знайдено — пропускаю")
            continue
        ws = wb[sheet_name]
        n = update_sheet(ws, rows, QTY_OVERRIDES.get(sheet_name))
        total_updated += n
        # підсумок з листа
        for r in range(ws.max_row, 0, -1):
            v = ws.cell(row=r, column=3).value
            if v and "ПІДСУМОК" in str(v):
                total_cell = ws.cell(row=r, column=11).value
                print(f"✅ {sheet_name}: оновлено {n} позицій. Підсумок: {total_cell}")
                break

    wb.save(XLSX)
    print(f"\n📊 Всього оновлено: {total_updated} позицій у {XLSX}")


if __name__ == "__main__":
    main()
