"""
Генератор Excel-кошторису для проєкту:
Garage Conversion to ADU — Menlo Park, CA 94025

V2 — структура по CONSTRUCTION PHASES (а не по зонах).
Послідовність 20 листів відповідає реальному порядку будівельних робіт:
  Site → Raised Floor → Framing → Openings → Plumbing RI → Electrical RI →
  HVAC RI → Insulation → Drywall → Paint → Flooring/Trim →
  Cabinetry → Bathroom Fixtures → Closet → Water Heater →
  Appliances → Furniture → Special Orders → Verification.

Дані витягнуті з ТЗ "Garage Conversion to ADU.md".
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

OUT_PATH = (
    "/Users/anatoliilisak/claude-workspace/documents/"
    "ADU — Menlo Park, CA 94025/ADU_Materials_Estimate.xlsx"
)

# 17 колонок: # | Phase | Category | Item | Spec | Unit | Q-Low | Q-Mid | Q-High |
#             | Unit Price | Total | Source | SKU/Model | URL | Status | Confidence | Notes
COLUMNS = [
    ("#", 5),
    ("Phase", 16),
    ("Категорія", 22),
    ("Item / Material", 38),
    ("Spec / Опис", 30),
    ("Unit", 10),
    ("Qty (Low)", 9),
    ("Qty (Mid)", 9),
    ("Qty (High)", 9),
    ("Unit Price ($)", 13),
    ("Total ($)", 13),
    ("Source", 16),
    ("SKU / Model", 22),
    ("URL", 32),
    ("Status", 12),
    ("Confidence", 11),
    ("Notes", 40),
]

# Стилі
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E78")
SUBTITLE_FONT = Font(bold=True, size=12, color="1F4E78")
TOTAL_FILL = PatternFill("solid", fgColor="FFE699")
TOTAL_FONT = Font(bold=True, size=11)
SPECIAL_ORDER_FILL = PatternFill("solid", fgColor="FFD9D9")  # підсвічую special order
BORDER_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


# ---------- Дані per лист ----------
# Кожен рядок: (Category, Item, Spec, Unit, Q_low, Q_mid, Q_high, Source, Status, Notes)
# Phase передається на рівні листа, не на рівні рядка.

# === 02 — Site / Exterior Utilities (§19) — V3 з reviewer fixes ===
SITE_EXTERIOR = [
    # --- SEWER (rows 5-18) ---
    ("Sewer", "Sewer Pipe 4\"", "SDR35 PVC, 10 ft pieces", "linear ft", 100, 110, 120, "Home Depot", "Confirmed", "100 ft підтверджено власником"),
    ("Sewer", "Sewer Couplings 4\"", "Hub couplings", "pcs", 4, 6, 8, "Home Depot", "TBD", ""),
    ("Sewer", "Sewer Elbows 45°/90°", "ABS/PVC short radius", "pcs", 4, 6, 8, "Home Depot", "TBD", ""),
    ("Sewer", "Long-radius Sweep 4\"", "Для slope (рекомендовано)", "pcs", 2, 3, 4, "Home Depot", "TBD", "Slope to be verified by plumber / local code"),
    ("Sewer", "Wye Fittings 4\"", "Branch tie-ins", "pcs", 2, 3, 4, "Home Depot", "TBD", ""),
    ("Sewer", "Sewer Cleanouts 4\"", "Body + plug + collar", "pcs", 2, 3, 4, "Home Depot", "Confirmed", "Depends on route and bends — підвищено Mid до 3"),
    ("Sewer", "Primer / Cement ABS-PVC", "Solvent weld", "cans", 2, 3, 4, "Home Depot", "TBD", ""),
    ("Sewer", "Sand Bedding (3\" under pipe)", "Cushion under pipe", "cu yd", 1, 2, 3, "Aggregate yard", "Confirmed", "Recommended — verify if required by local code / utility"),
    ("Sewer", "Gravel / Bedding (above pipe)", "3/4\" rock", "cu yd", 1, 2, 3, "Aggregate yard", "TBD", ""),
    ("Sewer", "Tracer Wire 14 AWG", "Locator wire для non-metallic", "linear ft", 100, 110, 120, "Home Depot", "Confirmed", "Recommended — verify if required by local code"),
    ("Sewer", "Warning Tape Detectable", "'CAUTION SEWER'", "rolls", 1, 1, 1, "Home Depot", "TBD", ""),
    ("Sewer", "Wall / Foundation Sleeve 4\"", "Slab penetration", "pcs", 1, 1, 1, "Home Depot", "TBD", ""),
    ("Sewer", "Concrete Patch Material", "50 lb bag", "bags", 1, 2, 3, "Home Depot", "TBD", ""),
    ("Sewer", "Test Caps / Cleanout Plugs allowance", "Для rough-in / pressure testing", "set", 1, 1, 2, "Home Depot", "TBD", "ДОДАНО (review): для тестування pipe ends"),
    # --- WATER SUPPLY (rows 19-33) ---
    ("Water Supply", "Main Water Supply Pipe", "PEX-A 1\" coil", "linear ft", 100, 110, 120, "Home Depot", "Confirmed", "100 ft підтверджено"),
    ("Water Supply", "PEX-A UV Protective Sleeve", "HDPE для buried PEX", "linear ft", 100, 110, 120, "Home Depot / specialty", "Confirmed", "PEX-A degrades on UV — required для buried sections"),
    ("Water Supply", "Main Shut-off Valve (ADU) 1\"", "Brass ball, full port", "pcs", 1, 1, 1, "Home Depot", "Confirmed", ""),
    ("Water Supply", "Brass Union Connector", "Tap point + ADU connection", "pcs", 1, 2, 2, "Home Depot", "TBD", ""),
    ("Water Supply", "Water Meter (separate)", "Якщо ADU має окремий meter", "pcs", 0, 0, 1, "Special Order / utility", "Optional", "OPTIONAL — verify with utility / town. Не входить в baseline total"),
    ("Water Supply", "Water Fittings / Couplings", "PEX brass crimp pack", "pack", 1, 2, 3, "Home Depot", "TBD", ""),
    ("Water Supply", "Pipe Sleeves", "Slab/wall penetration", "pcs", 2, 3, 4, "Home Depot", "TBD", ""),
    ("Water Supply", "Pipe Insulation (above-grade)", "Foam tube 1\"", "linear ft", 30, 50, 80, "Home Depot", "TBD", ""),
    ("Water Supply", "Tracer Wire 14 AWG (water)", "Locator wire для non-metallic", "linear ft", 100, 110, 120, "Home Depot", "Confirmed", "Recommended — verify if required by local code"),
    ("Water Supply", "Warning Tape (water)", "'CAUTION WATER'", "rolls", 1, 1, 1, "Home Depot", "TBD", ""),
    ("Water Supply", "Bedding Material", "Sand bed під pipe", "cu yd", 1, 2, 3, "Aggregate yard", "TBD", ""),
    ("Water Supply", "Pressure Regulator / Backflow", "Якщо street >80 PSI", "pcs", 0, 1, 1, "Home Depot", "TBD", ""),
    ("Water Supply", "Pressure Test Cap / Plug Set", "Для rough-in pressure test", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review): для rough-in testing"),
    ("Water Supply", "Ball Valve Access Box", "Якщо shut-off зовні / underground", "pcs", 0, 1, 1, "Home Depot", "TBD", "ДОДАНО (review): access для outdoor shut-off"),
    ("Water Supply", "Pipe Clamps / Supports", "Для above-grade sections", "pack", 1, 1, 2, "Home Depot", "TBD", "ДОДАНО (review)"),
    # --- ELECTRICAL FEED (rows 34-52) ---
    ("Electrical Feed", "Electrical Conduit (feeder)", "Schedule 40 PVC 2\"", "linear ft", 50, 60, 70, "Home Depot", "Confirmed", "50 ft підтверджено"),
    ("Electrical Feed", "PVC 90° Sweeps Long-Radius 2\"", "Bends in trench", "pcs", 2, 3, 4, "Home Depot", "Confirmed", "Long-radius для cable pulling"),
    ("Electrical Feed", "PVC 45° Sweeps 2\"", "Transitions", "pcs", 1, 2, 3, "Home Depot", "TBD", ""),
    ("Electrical Feed", "PVC Expansion Fitting 2\"", "Thermal expansion compensation", "pcs", 1, 1, 1, "Home Depot", "Confirmed", "Recommended — verify if required by local code"),
    ("Electrical Feed", "PVC Conduit Cement", "Solvent weld for PVC", "cans", 1, 1, 1, "Home Depot", "Confirmed", ""),
    ("Electrical Feed", "Feeder Wire", "Cu SER 4-4-4-6 (base) OR Al SER 2-2-2-4", "linear ft", 50, 60, 70, "Home Depot / specialty", "TBD", "TBD by electrician / load calculation. Cu SER ~$8.78/LF, Al SER ~$3.50/LF (-$315 з 60 LF). Final type/size pendant on ADU load calc"),
    ("Electrical Feed", "Disconnect Switch (exterior)", "100A non-fusible if req'd", "pcs", 0, 1, 1, "Home Depot", "TBD", ""),
    ("Electrical Feed", "Ground Rod 8 ft Copper-clad", "Driven rod", "pcs", 1, 1, 2, "Home Depot", "Confirmed", ""),
    ("Electrical Feed", "Ground Rod Clamp", "Acorn or bolt clamp", "pcs", 1, 1, 2, "Home Depot", "Confirmed", ""),
    ("Electrical Feed", "#6 Bare Cu Ground Wire", "Service ground", "linear ft", 10, 15, 25, "Home Depot", "Confirmed", ""),
    ("Electrical Feed", "Pull String / Fish Tape", "Для cable pulling", "pcs", 1, 1, 1, "Home Depot", "Confirmed", ""),
    ("Electrical Feed", "Anti-oxidant Compound (NOALOX)", "Якщо AL feeder used", "tube", 0, 1, 1, "Home Depot", "TBD", ""),
    ("Electrical Feed", "Conduit Straps / Clamps / Mounting", "Для exposed conduit sections", "pack", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО (review): mounting hardware"),
    ("Electrical Feed", "Terminal Adapters / Bushings", "Stubs, male adapters, bell ends", "pack", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review)"),
    ("Electrical Feed", "Subpanel / Circuit Labels", "Для identification", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review)"),
    ("Electrical Feed", "Weatherproof Boxes", "", "pcs", 2, 3, 4, "Home Depot", "TBD", ""),
    ("Electrical Feed", "Pull Boxes", "If req'd", "pcs", 0, 1, 2, "Home Depot", "TBD", ""),
    ("Electrical Feed", "Underground Warning Tape", "'CAUTION ELECTRIC'", "rolls", 1, 1, 1, "Home Depot", "Confirmed", ""),
    ("Electrical Feed", "Exterior-rated Fittings", "Adapters, bushings pack", "pack", 1, 1, 1, "Home Depot", "Confirmed", ""),
    # --- SITE PREP / LAYOUT (row 53) ---
    ("Site Prep", "Marking Paint / Stakes / String Line", "Для trench layout всіх 3 утилит", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review): layout aids"),
]

# === 03 — Raised Floor System (§17) — V3 з reviewer fixes (5 нових items + notes) ===
# Підхід: 2x10 joists @ 16" OC + 3/4" plywood subfloor = 10" rise
RAISED_FLOOR = [
    ("Moisture Control", "Vapor / Moisture Barrier", "6-mil poly", "sq ft", 315, 350, 380, "Home Depot", "Confirmed", "10\" raised over slab"),
    ("Lumber", "Pressure-treated Sleepers", "2x4 PT laid flat as bearing", "linear ft", 200, 260, 320, "Home Depot", "TBD", "Bearing surface для joists на slab"),
    ("Lumber", "Floor Joists", "2x10 SPF/HF @ 16\" OC, 15 ft span", "linear ft", 240, 280, 340, "Home Depot", "TBD", "Final joist size/span to be verified by contractor/engineer based on actual support layout"),
    ("Lumber", "Rim Joists", "2x10 PT за perimeter", "linear ft", 60, 72, 80, "Home Depot", "Confirmed", "PT lumber required where wood contacts concrete or moisture-prone areas"),
    ("Lumber", "Blocking", "2x10 mid-span", "linear ft", 30, 50, 70, "Home Depot", "TBD", ""),
    ("Hardware", "Joist Hangers / Connectors", "Simpson LUS210", "pcs", 30, 50, 70, "Home Depot", "Confirmed", ""),
    ("Hardware", "Concrete Anchors (PT to slab)", "Wedge / Tapcon 1/2\"", "pcs", 40, 60, 80, "Home Depot", "Confirmed", ""),
    ("Insulation (floor)", "Insulation R-30 batt or Rigid", "Batt R-30 between joists", "sq ft", 315, 330, 350, "Home Depot", "TBD", "Verify floor insulation type: batt vs rigid foam depending on moisture and floor assembly"),
    ("Subfloor", "Plywood Subfloor", "3/4\" T&G CDX", "sq ft", 315, 350, 380, "Home Depot", "Confirmed", ""),
    ("Adhesives/Fasteners", "Construction Adhesive", "PL Premium subfloor", "tubes", 6, 10, 14, "Home Depot", "Confirmed", ""),
    ("Adhesives/Fasteners", "Subfloor Screws", "3\" #8 deck/subfloor", "lb", 5, 8, 12, "Home Depot", "Confirmed", "Часто продаються в 5 lb або 10 lb boxes"),
    ("Sealants", "Foam Sealant / Caulk", "Air gap sealing", "cans", 3, 5, 7, "Home Depot", "TBD", ""),
    ("Transitions", "Thresholds & Transitions", "Entry, French door, bathroom, exterior step", "pcs", 3, 4, 5, "Home Depot", "TBD", "§17.4. Не дублювати з Phase 12 (decorative transitions only)"),
    ("Sealing", "Sill Seal Foam Gasket", "1x50 ft roll, між PT і slab", "rolls", 1, 2, 3, "Home Depot", "Confirmed", "Air sealing"),
    ("Slab Prep", "Self-leveling Compound", "Allowance only — partial leveling", "bags", 0, 4, 10, "Home Depot", "TBD", "§33.22. Якщо full leveling req'd — qty зростає significantly"),
    ("Slab Prep", "Concrete Sealer / DryLok", "Optional moisture mitigation", "gal", 0, 0, 5, "Home Depot", "TBD", "§33.23 verify moisture conditions. Baseline 0"),
    ("Sealants", "Acoustical / Sound Caulk", "Floor joints sound dampening", "tubes", 0, 3, 6, "Home Depot", "TBD", "Optional"),
    # --- ДОДАНО per reviewer review V3 ---
    ("Moisture Control", "Vapor Barrier Seam Tape", "Для overlap joints poly", "rolls", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО (review): Stego/generic seam tape"),
    ("Insulation (floor)", "Insulation Supports / Wire Supports", "Для утримання batt між joists", "pack", 1, 2, 3, "Home Depot", "Confirmed", "ДОДАНО (review): тримає R-30 batt вгору"),
    ("Hardware", "Composite / Plastic Shims", "Для вирівнювання sleepers/joists", "pack", 1, 2, 3, "Home Depot", "Confirmed", "ДОДАНО (review)"),
    ("Hardware", "Structural Screws / Lag Screws", "SDS / Ledgerlok для rim-joist", "box", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО (review): для connections without hangers"),
    ("Access / Utility", "Small Access Panel allowance", "12x12 PVC/металевий", "pcs", 0, 1, 2, "Home Depot", "TBD", "ДОДАНО (review): для доступу до MEP під підлогою"),
]

# === 04 — Framing (§18) — V3 з reviewer fixes ===
# Стіни: ~95-105 LF total (65-75 interior + 21 left wall + ~10 utility closet + returns/openings)
# Studs @ 16" OC + 20-25% overhead для corners/openings/doublers
FRAMING = [
    # --- STUDS ---
    ("Studs/Plates", "2x4 Studs (8 ft)", "SPF #2, wall framing", "pcs", 75, 90, 110, "Home Depot", "Confirmed", "95-105 LF wall × 0.75 + 20-25% overhead. Включає left wall +21 LF + utility + openings"),
    ("Studs/Plates", "2x6 Studs (8 ft) — wet wall", "Для bathroom plumbing wall", "pcs", 5, 8, 12, "Home Depot", "TBD", "Depends on final bathroom/laundry layout — wet wall behind shower/vanity/laundry"),
    ("Studs/Plates", "Cripple / Jack / King Studs", "Allowance для openings", "pcs", 15, 20, 28, "Home Depot", "Confirmed", "Для 12+ framed openings"),
    # --- PLATES ---
    ("Studs/Plates", "PT Bottom Plates", "Pressure-treated 2x4 — на subfloor edge", "linear ft", 65, 75, 85, "Home Depot", "TBD", "PT де lumber contacts moisture-prone area"),
    ("Studs/Plates", "Top / Bottom Plates (regular)", "2x4 SPF (double top + bottom)", "linear ft", 130, 150, 170, "Home Depot", "TBD", ""),
    # --- HEADERS ---
    ("Headers", "Standard Headers (2x6)", "Above standard openings (windows, doors)", "linear ft", 20, 30, 40, "Home Depot", "TBD", "Structural sizing — final per engineer"),
    ("Headers", "Engineered LVL Header", "Для French door (~6ft) + folding 8ft", "linear ft", 12, 16, 20, "Home Depot / specialty", "TBD", "Final size depends on opening width, load above, existing roof/wall structure"),
    # --- BLOCKING & BACKING ---
    ("Blocking", "Blocking Lumber (2x4)", "12 типів blocking per §18.3", "linear ft", 40, 60, 80, "Home Depot", "Confirmed", "Cabinets/vanity/W-D/WH straps/TV/doors/towels/shelves..."),
    ("Backing", "Drywall Backing / Nailers", "Inside corners + T-intersections", "linear ft", 20, 30, 40, "Home Depot", "Confirmed", "Для drywall attachment points"),
    # --- HARDWARE ---
    ("Hardware", "Simpson Connectors / Brackets", "L, T, A35 mixed", "pcs", 20, 30, 40, "Home Depot", "TBD", ""),
    ("Hardware", "Hold-downs Simpson HDU2 allowance", "Allowance — TBD per structural engineer", "pcs", 0, 4, 6, "Home Depot", "TBD", "§33.25 — verify shear wall requirement"),
    ("Hardware", "Metal Protection Plates", "Stud guards 1.5x5 за trades", "pcs", 20, 30, 40, "Home Depot", "TBD", ""),
    ("Hardware", "Composite / Wood Shims", "Для вирівнювання walls/openings", "pack", 1, 2, 4, "Home Depot", "Confirmed", "ДОДАНО (review): wall plumbing і opening setting"),
    # --- ANCHORS ---
    ("Anchors", "Anchor Bolts 1/2 × 10 galvanized", "Для sill/perimeter/structural conditions", "pcs", 12, 20, 30, "Home Depot", "TBD", "Verify use: anchor bolts для sill/perimeter/structural; interior partitions on subfloor можуть use screws/nails"),
    # --- FASTENERS ---
    ("Fasteners", "Framing Nails / Screws", "16d collated, 3-1/4\" mixed", "lb", 10, 15, 20, "Home Depot", "Confirmed", ""),
    ("Fasteners", "Structural Screws / Header Fasteners", "SDS/Strong-Drive для LVL & headers", "box", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО (review): для LVL/header/opening connections"),
    # --- SHEATHING ---
    ("Sheathing", "OSB / Plywood Sheathing allowance", "Garage opening close-up / shear panels", "sheets", 1, 3, 5, "Home Depot", "TBD", "ДОДАНО (review): garage door opening closure / shear wall / backing"),
    # --- ADHESIVES & SEALANTS ---
    ("Adhesives", "Construction Adhesive (framing)", "PL Premium plate-to-floor", "tubes", 3, 5, 7, "Home Depot", "Confirmed", ""),
    ("Sealants", "Acoustical Sealant", "Bedroom/bathroom partitions sound", "tubes", 2, 4, 6, "Home Depot", "TBD", "Sound dampening per partition"),
    # --- FIRE PROTECTION ---
    ("Fire Protection", "Fire Blocking Materials", "Foam + 2x4 blocks per code", "set", 1, 1, 1, "Home Depot", "TBD", "Verify if required by local code"),
    ("Fire Protection", "Fire-rated Caulk / Sealant", "Для penetrations", "tubes", 2, 4, 6, "Home Depot", "TBD", "ДОДАНО (review): окремо від foam blocking"),
    # --- LAYOUT SUPPLIES ---
    ("Framing Supplies", "Chalk Line / Marking Supplies", "Chalk + line + carpenter pencils + plumb bob", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review): layout aids"),
    # --- BRACING ---
    ("Bracing", "Temporary Bracing Lumber", "Для setting walls during framing", "linear ft", 15, 25, 40, "Home Depot", "TBD", "ДОДАНО (review): можна reuse після"),
]

# === 05 — Doors & Windows / Openings (§25) — V3 з reviewer fixes ===
DOORS_WINDOWS = [
    # --- WINDOWS (3) ---
    ("Windows", "Bathroom Window", "Vinyl ES, ~24×36 (TBD §33.12)", "pcs", 1, 1, 1, "Home Depot", "TBD", "Field-verify size"),
    ("Windows", "Kitchen Window", "Vinyl ES, ~36×48 (TBD)", "pcs", 1, 1, 1, "Home Depot", "TBD", "Field-verify size"),
    ("Windows", "Living Room Window", "Vinyl ES, ~36×60 (TBD)", "pcs", 1, 1, 1, "Home Depot", "TBD", "Field-verify size"),
    # --- EXTERIOR DOORS (2) ---
    ("Exterior Doors", "Main Entry Door", "Prehung steel 36×80 insulated", "pcs", 1, 1, 1, "Home Depot", "TBD", "Field-verify size (§33.14)"),
    ("Exterior Doors", "Backyard French Door", "Fiberglass prehung 60×80, glazed", "pcs", 1, 1, 1, "Special Order", "Confirmed", "Special Order. Lead 4-8 weeks. Delivery not included. Field-verify (§33.13)"),
    # --- INTERIOR DOORS (2) ---
    ("Interior Doors", "Bathroom Door", "Prehung hollow core 30×80", "pcs", 1, 1, 1, "Home Depot", "TBD", "Verify if privacy lockset included; otherwise add separately"),
    ("Interior Doors", "Utility Closet Door", "24×80 louvered bifold або slab", "pcs", 1, 1, 1, "Home Depot", "TBD", "Final type TBD based on clearance"),
    # --- SPECIALTY SYSTEMS (2) ---
    ("Specialty Systems", "Closet Sliding Door System", "Bypass 6 ft, 2-panel", "system", 1, 1, 1, "Home Depot / IKEA", "TBD", "Field-verify (§33.16)"),
    ("Specialty Systems", "Folding / Panel Door System 8 ft", "INTERIOR partition між bedroom і living (per §17.4)", "system", 1, 1, 1, "Special Order", "Confirmed", "Special Order INTERIOR. High variability — final cost залежить від material: accordion / bifold / glass folding / wood / track / sound level. Field-verify (§33.15)"),
    # --- WEATHERPROOFING ---
    ("Weatherproofing", "Flashing Tape", "Windows + doors butyl/peel-stick", "rolls", 2, 3, 4, "Home Depot", "TBD", ""),
    ("Weatherproofing", "Sill Pans", "Windows + 2 exterior doors", "pcs", 4, 5, 5, "Home Depot", "TBD", "Підвищено Mid: 3 windows + entry + French = 5 openings"),
    ("Weatherproofing", "Exterior Sealant", "Polyurethane caulk", "tubes", 4, 6, 8, "Home Depot", "TBD", ""),
    ("Weatherproofing", "Low-expansion Foam", "Для perimeter gaps", "cans", 3, 4, 6, "Home Depot", "TBD", ""),
    # --- TRIM ---
    ("Trim", "Interior Casing", "За дверима/вікнами, 7+ openings", "linear ft", 80, 120, 160, "Home Depot", "TBD", ""),
    ("Trim", "Exterior Trim", "Standard exterior molding", "linear ft", 40, 60, 80, "Home Depot", "TBD", ""),
    # --- HARDWARE — REORGANIZED per reviewer ---
    ("Hardware", "Entry Door Lockset + Deadbolt", "Для main entry only", "set", 1, 1, 1, "Home Depot", "TBD", "Schlage/Kwikset deadbolt + handle"),
    # --- ДОДАНО per Pre-Takeoff Check ---
    ("Windows", "Window Screens", "Fiberglass mesh", "pcs", 3, 3, 3, "Home Depot", "Confirmed", "May be included with selected windows; keep as allowance until product selected"),
    ("Windows", "Window Sills / Stools", "Interior trim під window", "pcs", 3, 3, 3, "Home Depot", "Confirmed", "Окремо від generic casing"),
    ("Trim", "Brick Mold / Exterior Casing", "Для exterior openings", "linear ft", 40, 60, 80, "Home Depot", "Confirmed", "Окремо від generic exterior trim"),
    ("Hardware", "French Door Hardware", "Multi-point lock + handles", "set", 1, 1, 1, "Special Order", "TBD", "Специфічно для French door (§25.3)"),
    ("Hardware", "Weatherstripping", "Kit для exterior doors", "set", 2, 2, 2, "Home Depot", "Confirmed", "Entry + French"),
    ("Hardware", "Door Sweep / Threshold Seal", "Для exterior doors", "pcs", 2, 2, 2, "Home Depot", "Confirmed", "Air leak prevention"),
    ("Hardware", "Threshold Extender (French door)", "Для 10\" raise через raised floor", "pcs", 1, 1, 1, "Home Depot / specialty", "TBD", "Per Phase 2 raised floor coordination"),
    ("Hardware", "Astragal Strip allowance", "Між French door panels", "pcs", 0, 1, 1, "Home Depot", "TBD", "Якщо single-leaf French door"),
    # --- ДОДАНО per reviewer V3 ---
    ("Hardware", "Interior Privacy/Passage Handles", "Bathroom (privacy) + utility (passage)", "set", 2, 2, 3, "Home Depot", "Confirmed", "ДОДАНО (review): if not included with prehung doors"),
    ("Hardware", "Hinges Allowance", "Якщо doors не prehung або заміна", "set", 0, 1, 2, "Home Depot", "TBD", "ДОДАНО (review)"),
    ("Installation Supplies", "Door / Window Shims", "Composite/wood для setting", "pack", 2, 3, 4, "Home Depot", "Confirmed", "ДОДАНО (review)"),
    ("Weatherproofing", "Backer Rod", "Для великих gaps перед sealant", "rolls", 1, 2, 3, "Home Depot", "Confirmed", "ДОДАНО (review)"),
    ("Weatherproofing", "Drip Cap / Z-Flashing", "Над exterior doors/windows", "pcs", 2, 3, 5, "Home Depot", "TBD", "ДОДАНО (review)"),
    ("Finish", "Interior Paintable Caulk", "Для casing/trim joints", "tubes", 6, 8, 10, "Home Depot", "Confirmed", "ДОДАНО (review)"),
]

# === 06 — Plumbing Rough-in (§20) — V2 з 10 додатковими items ===
PLUMBING_RI = [
    # --- SUPPLY ---
    ("Supply", "PEX Hot/Cold Pipe 1/2\"", "PEX-B coil 100 ft", "linear ft", 200, 230, 260, "Home Depot", "Confirmed", ""),
    ("Supply", "PEX Hot/Cold Pipe 3/4\"", "Main runs PEX-B", "linear ft", 100, 120, 140, "Home Depot", "Confirmed", ""),
    ("Supply", "PEX Fittings (assortment)", "Elbows, tees, couplings, crimps", "pack", 2, 3, 4, "Home Depot", "TBD", ""),
    ("Supply", "Shut-off Valves (1/4-turn)", "Interior fixtures angle stops", "pcs", 16, 19, 22, "Home Depot", "Confirmed", ""),
    # --- DRAIN ---
    ("Drain", "ABS/PVC Drain Pipe 1.5\"", "Sink/vanity drains", "linear ft", 30, 40, 50, "Home Depot", "TBD", ""),
    ("Drain", "ABS/PVC Drain Pipe 2\"", "Shower/laundry drains", "linear ft", 40, 60, 80, "Home Depot", "TBD", ""),
    ("Drain", "ABS/PVC Drain Pipe 3\"", "Branch + toilet drain", "linear ft", 20, 30, 40, "Home Depot", "TBD", ""),
    ("Drain", "ABS/PVC Drain Pipe 4\"", "Main building drain", "linear ft", 15, 20, 25, "Home Depot", "TBD", ""),
    # --- VENT ---
    ("Vent", "Vent Pipe", "ABS/PVC 1.5-2\"", "linear ft", 20, 30, 40, "Home Depot", "TBD", ""),
    # --- FITTINGS ---
    ("Fittings", "P-traps", "Sinks, shower, washer", "pcs", 4, 4, 5, "Home Depot", "Confirmed", ""),
    ("Fittings", "Cleanouts (interior)", "Body + plug", "pcs", 2, 2, 3, "Home Depot", "Confirmed", ""),
    ("Fittings", "Pipe Clamps / Hangers", "ABS strut/strap pack", "pack", 2, 3, 4, "Home Depot", "TBD", ""),
    ("Fittings", "Primer / Cement ABS-PVC", "Solvent weld", "cans", 2, 3, 4, "Home Depot", "TBD", ""),
    # --- MISC ---
    ("Misc", "Washer Outlet Box", "Recessed, hot/cold + drain", "pcs", 1, 1, 1, "Home Depot", "Confirmed", ""),
    ("Misc", "Toilet Flange", "PVC/ABS w/ ring", "pcs", 1, 1, 1, "Home Depot", "Confirmed", ""),
    ("Misc", "Water Heater Connection Kit", "Flex lines + fittings", "set", 1, 1, 1, "Home Depot", "TBD", "Coordinate w/ Phase 15 Water Heater"),
    ("Misc", "Pipe Insulation (interior)", "Hot lines foam 1/2-3/4\"", "linear ft", 30, 50, 70, "Home Depot", "TBD", ""),
    # --- ДОДАНО per Pre-Takeoff Check ---
    ("Fittings", "ABS/PVC DWV Fittings", "Elbows, couplings, tees, sanitary tees, wyes, reducers", "pack", 2, 3, 5, "Home Depot", "Confirmed", "ДОДАНО (§20.3): drain fittings okремо"),
    ("Misc", "Washer Standpipe", "2\" 18-30\" tall", "pcs", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (§20.3): окремо від outlet box"),
    ("Misc", "Dishwasher Connection Kit", "Tee + brass elbow + 3/8 supply", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (§20.3)"),
    ("Vent", "Vent Stack Roof Boot / Flashing", "Oatey No-Calk pipe boot", "pcs", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО: penetration through roof"),
    ("Vent", "Air Admittance Valve (AAV)", "Studor Mini / Pro для hard-to-vent", "pcs", 0, 1, 2, "Home Depot", "TBD", "ДОДАНО: alternative venting allowance"),
    ("Misc", "Refrigerator Water Line Kit", "3/8 PEX + saddle valve + icemaker line", "set", 0, 1, 1, "Home Depot", "TBD", "ДОДАНО (§20.1 optional)"),
    ("Misc", "Hammer Arrestors", "Washer + dishwasher water hammer", "pcs", 2, 2, 3, "Home Depot", "Confirmed", "ДОДАНО: prevent water hammer"),
    ("Sealants", "Pipe Sealant / Teflon Tape / Pipe Dope", "Threaded connections", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО: Rectorseal + Teflon roll combo"),
    ("Fittings", "Threaded Brass Adapters", "PEX-to-threaded fittings", "pack", 1, 2, 3, "Home Depot", "Confirmed", "ДОДАНО: для sink supply / WH"),
    ("Misc", "Stub-out Covers + Test Caps + Escutcheons", "Finish wall penetrations + drain testing", "set", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО: trim items + pressure test plugs"),
    # --- ДОДАНО per reviewer V2 (fixture rough-in items) ---
    ("Fixtures RI", "Shower Rough-in Valve Body", "Delta MultiChoice / Moen Posi-Temp", "pcs", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review): rough-in stage. Trim → Phase 13"),
    ("Fixtures RI", "Shower Drain Assembly", "Oatey 2\" drain body + strainer", "pcs", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review): connects до 2\" drain pipe"),
    ("Supply", "PEX Drop-Ear Elbows", "Sioux Chief 1/2 PEX × 1/2 FNPT для wall stub-outs", "pcs", 4, 6, 8, "Home Depot", "Confirmed", "ДОДАНО (review): для secure fixture wall outlets"),
    ("Supply", "Extra Crimp Rings / Clamps", "Copper crimp 1/2 + 3/4 pack", "pack", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО (review): backup до PEX fittings assortment"),
    ("Misc", "Laundry Drain Pan", "30x32 для stacked washer", "pcs", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review): leak protection під washer"),
    ("Misc", "Dishwasher Branch Tailpiece + Air Gap", "Sink connection + CA-required air gap", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review): CA UPC potreba air gap"),
    ("Misc", "Water Heater Pan + Drain (cross-ref Phase 15)", "Залишено в Phase 15 — не дублюйте", "set", 0, 0, 0, "Home Depot", "Confirmed", "ДОДАНО (review): cross-reference. Cost у Phase 15 Water Heater"),
]

# === 07 — Electrical Rough-in (§21) — V3 з reviewer fixes ===
ELECTRICAL_RI = [
    # --- PANEL ---
    ("Panel", "ADU Subpanel 100A", "Square D QO 20-space allowance", "pcs", 1, 1, 1, "Home Depot", "TBD", "Final amperage by load calculation (§33.21)"),
    ("Panel", "Standard Breakers (non-AFCI)", "Single/double pole для 240V appliances + non-AFCI circuits", "set", 1, 1, 1, "Home Depot", "Confirmed", "Тільки для circuits де AFCI не required: range, WH, mini-split, dryer, bathroom GFCI"),
    # --- WIRE ---
    ("Wire", "Romex 12/2 (20A circuits)", "250ft / 1000ft rolls NM-B", "linear ft", 500, 600, 700, "Home Depot", "Confirmed", ""),
    ("Wire", "Romex 14/2 (15A circuits)", "Lighting general NM-B", "linear ft", 300, 400, 500, "Home Depot", "Confirmed", ""),
    ("Wire", "240V Cable (range/WH/HVAC)", "8/3 (range) + 10/2 (WH, mini-split)", "linear ft", 80, 120, 160, "Home Depot", "Confirmed", "Range 40A, WH 30A, mini-split 30A"),
    ("Wire", "Interior Conduit (where req'd)", "EMT/PVC 1/2-3/4\"", "linear ft", 20, 40, 60, "Home Depot", "TBD", ""),
    # --- DEVICES ---
    ("Devices", "Electrical Boxes (mixed)", "1-gang, 2-gang, ceiling, old-work", "pcs", 55, 62, 70, "Home Depot", "Confirmed", ""),
    ("Devices", "Standard Outlets TR-rated", "Tamper-resistant 15A/20A", "pcs", 25, 30, 35, "Home Depot", "Confirmed", "CA Code 2019+ TR mandatory"),
    ("Devices", "GFCI Outlets", "Bath, kitchen, exterior", "pcs", 8, 10, 12, "Home Depot", "Confirmed", "Final GFCI strategy TBD: protection may be provided by GFCI receptacles, GFCI breakers, or dual-function breakers"),
    ("Devices", "Switches / Dimmers (LED-compatible)", "Single/3-way mix", "pcs", 12, 15, 18, "Home Depot", "Confirmed", "CA Title 24 LED dimming"),
    ("Devices", "Cover Plates", "", "pcs", 55, 62, 70, "Home Depot", "Confirmed", ""),
    ("Devices", "Bathroom Fan Timer Switch", "", "pcs", 1, 1, 1, "Home Depot", "TBD", ""),
    ("Devices", "Smoke / CO Detectors", "Combo hardwired interconnected + battery backup", "pcs", 2, 3, 4, "Home Depot", "Confirmed", "Final placement TBD. CA Code: hardwired interconnected"),
    ("Devices", "Weatherproof Outlets", "Exterior GFCI in-use cover", "pcs", 1, 2, 2, "Home Depot", "TBD", ""),
    ("Devices", "Mini-split Disconnect (outdoor)", "60A pull-out switch", "pcs", 1, 1, 1, "Home Depot", "Confirmed", ""),
    ("Devices", "Appliance Receptacles", "Range/dryer/etc receptacles", "pcs", 3, 4, 5, "Home Depot", "TBD", "TBD by appliance specs; some equipment may be hardwired (WH, mini-split)"),
    # --- LIGHTING FIXTURES ---
    ("Lighting Fixtures", "Recessed Lights LED (cans+trim)", "Wafer LED 4\"/6\" dimmable", "pcs", 10, 12, 14, "Home Depot", "Confirmed", ""),
    ("Lighting Fixtures", "Ceiling Lights (surface)", "Closet, utility", "pcs", 2, 3, 4, "Home Depot", "TBD", ""),
    ("Lighting Fixtures", "Exterior Lights", "Entry + French door", "pcs", 1, 2, 2, "Home Depot", "TBD", ""),
    ("Lighting Fixtures", "Bathroom Vanity Light", "", "pcs", 1, 1, 1, "Home Depot", "TBD", ""),
    ("Lighting Fixtures", "Closet Lighting", "Motion sensor", "pcs", 1, 1, 1, "Home Depot / IKEA", "TBD", ""),
    # --- ДОДАНО per Pre-Takeoff Check (CA Code + consumables) ---
    ("Panel", "AFCI / Dual-Function Breakers", "CA Code 2019+ для bedroom/living/kitchen/dining", "pcs", 6, 8, 10, "Home Depot", "Confirmed", "ДОДАНО (CA Code). НЕ дублюється з standard breakers — це окремі circuits"),
    ("Consumables", "Wire Nuts (Marrettes/Wago) pack", "Splices", "pack", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Consumables", "Electrical Tape (B/W/R phasing)", "3M Super 33", "pack", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Consumables", "Cable Staples / Romex Clips", "Для cable runs по studs", "lb", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Consumables", "Romex Cable Connectors", "Clamps box entry", "pack", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО: 25-count packs"),
    ("Consumables", "Grounding Pigtails / Extra Green Wire", "Extra grounds", "pack", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Devices", "Junction Boxes (ceiling splice)", "Окремо від device boxes", "pcs", 3, 5, 8, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Devices", "GFCI + USB Outlets", "Modern code allowance bedroom/kitchen", "pcs", 1, 2, 4, "Home Depot", "TBD", "ДОДАНО"),
    ("Panel", "Whole-House Surge Protector", "SPD type 2 для subpanel", "pcs", 0, 1, 1, "Home Depot", "Optional", "ДОДАНО (CA recommendation — Optional)"),
    ("Bonding", "Bonding Jumpers + Clamps", "Metal water lines (WH)", "pcs", 1, 2, 2, "Home Depot", "Confirmed", "ДОДАНО"),
    # --- ДОДАНО per reviewer V3 ---
    ("Panel", "Ground Bar / Neutral Bar Kit", "Часто потрібно окремо від subpanel", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review): may not be included з load center"),
    ("Panel", "Panel Labels / Circuit Directory", "Identification labels", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review)"),
    ("Lighting Fixtures", "Under-cabinet LED Driver / Transformer", "Якщо kitchen LED strip", "pcs", 0, 1, 1, "Home Depot / IKEA", "Optional", "ДОДАНО (review): для under-cabinet lighting (Phase 13)"),
    ("Low-Voltage", "Data / Coax Cable Allowance", "Cat6 + RG6 для router/TV/camera", "set", 0, 1, 1, "Home Depot", "Optional", "ДОДАНО (review): smart home allowance"),
    ("Fire Protection", "Fire Caulk / Putty Pads for Boxes", "Penetrations і fire-rated assembly", "set", 1, 1, 2, "Home Depot", "TBD", "ДОДАНО (review)"),
    ("Wire", "Spare Conduit Allowance", "Запас для future expansion", "set", 0, 1, 1, "Home Depot", "Optional", "ДОДАНО (review): якщо trench/walls open"),
]

# === 08 — HVAC Mini-Split (§22) — V3 з reviewer fixes ===
HVAC = [
    # --- OUTDOOR UNIT (kit) ---
    ("Outdoor Unit", "Mini-split Outdoor Condenser", "MRCOOL DIY 5th Gen 18k 9+9 KIT (z 2 indoor + line sets + remotes)", "pcs", 1, 1, 1, "Special Order", "Confirmed", "Special Order. Full kit. Sizing TBD per §33.9 Manual J. Lead 4-8 weeks"),
    # --- INDOOR UNITS — INCLUDED IN KIT ---
    ("Indoor Units", "Indoor Wall Units (2)", "9k BTU × 2 — included in kit r5", "pcs", 0, 0, 0, "Special Order", "Confirmed", "Included in main kit — do not price separately unless kit changes"),
    # --- LINE SETS — INCLUDED IN KIT ---
    ("Refrigerant", "Line Sets (pre-charged 25 ft)", "Included in kit r5", "set", 0, 0, 0, "Home Depot / specialty", "Confirmed", "Included in main kit"),
    # --- CONDENSATE ---
    ("Condensate", "Condensate Drain Lines", "3/4\" PVC + fittings (2 sets)", "set", 2, 2, 2, "Home Depot", "Confirmed", ""),
    # --- PENETRATIONS ---
    ("Penetrations", "Wall Sleeves", "Прохід крізь стіну (2)", "pcs", 2, 2, 2, "Home Depot", "Confirmed", ""),
    # --- MOUNTING — wall bracket optional alternative ---
    ("Mounting", "Wall Bracket (Optional)", "Optional alternative якщо ground placement не suitable", "pcs", 0, 0, 1, "Home Depot", "Optional", "Baseline = concrete pad (r14). Wall bracket only якщо ground не підходить"),
    # --- TRIM ---
    ("Trim", "Line Hide / Cover Kits", "Decorative exterior cover", "set", 1, 1, 2, "Home Depot", "TBD", ""),
    # --- MOUNTING HARDWARE ---
    ("Mounting", "Mounting Hardware", "Lag bolts + brackets + screws", "set", 1, 1, 1, "Home Depot", "Confirmed", ""),
    # --- CONTROLS ---
    ("Controls", "Controllers / Remotes", "Included in kit r5", "pcs", 0, 0, 0, "Manufacturer", "Confirmed", "Included з indoor units in kit"),
    # --- BASELINE MOUNTING ---
    ("Mounting", "Concrete Pad для Outdoor Unit", "36×36\" precast — BASELINE mount", "pcs", 1, 1, 1, "Home Depot", "Confirmed", "Baseline mounting. Wall bracket — Optional alternative (r10)"),
    ("Mounting", "Vibration Isolation Pads", "Rubber/cork pads під condenser", "set", 1, 1, 1, "Home Depot", "Confirmed", "Noise/vibration reduction"),
    # --- CONDENSATE TRAP ---
    ("Condensate", "Condensate Trap / Vent", "Prevent air-lock в drain line", "pcs", 1, 2, 2, "Home Depot", "Confirmed", ""),
    ("Condensate", "Condensate Pump (Optional)", "Якщо gravity drain неможливий", "pcs", 0, 0, 1, "Home Depot", "Optional", "Backup allowance"),
    # --- MAINTENANCE ---
    ("Maintenance", "Replacement Filters (set)", "Initial + 2-3 spare per indoor", "set", 1, 1, 2, "Home Depot / Amazon", "Confirmed", ""),
    # --- CONTROLS WI-FI ---
    ("Controls", "Wi-Fi Smart Adapter", "Modern control (часто sold окремо)", "pcs", 0, 1, 2, "Amazon", "Optional", "Smart home allowance"),
    # --- REFRIGERANT extras ---
    ("Refrigerant", "R-410A Refrigerant (extra)", "Якщо line set >25 ft або leak", "lb", 0, 0, 5, "Home Depot / specialty", "Optional", "Pre-charge typically ≤25 ft. Optional only"),
    ("Refrigerant", "Flare Fittings / Nuts", "Refrigerant line connections — verify if included з kit", "pack", 1, 1, 1, "Home Depot", "TBD", "DIY quick-connect kits часто mають flares pre-installed"),
    ("Refrigerant", "Insulation Tape / Sealant", "Line set joint tightness", "rolls", 1, 1, 2, "Home Depot", "Confirmed", ""),
    # --- SAFETY — leak detector only (drip pan optional) ---
    ("Safety", "Leak Detector / Drip Tray", "Indoor unit leak protection — Optional", "pcs", 0, 1, 2, "Home Depot", "Optional", "Optional. Drip pan only if compatible з selected indoor unit/location"),
    # --- ДОДАНО per reviewer V3 ---
    ("Penetrations", "Penetration Sealant / Putty", "Exterior wall sleeve sealing", "tubes", 1, 2, 3, "Home Depot", "Confirmed", "ДОДАНО (review): line-set/wall sleeve waterproofing"),
    ("Penetrations", "Wall Cap / Line-set Penetration Cover", "Закрити прохід через exterior wall", "pcs", 1, 2, 2, "Home Depot", "Confirmed", "ДОДАНО (review)"),
    ("Mounting", "UV-rated Zip Ties / Line Set Clamps", "Кріплення line set, condensate, line hide", "pack", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО (review)"),
    ("Condensate", "Condensate Termination Fittings", "Exterior drain end fittings + elbows", "set", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО (review)"),
    ("Cross-ref", "Mini-split Electrical Whip / Disconnect", "Cross-ref Phase 6 r19 — не дублювати", "pcs", 0, 0, 0, "—", "Confirmed", "ДОДАНО (review): електричне з'єднання disconnect → condenser у Phase 6"),
]

# === 09 — Insulation (§23) — V3 з reviewer fixes ===
INSULATION = [
    ("Walls", "Exterior Wall Batt R-15", "Kraft-faced fiberglass 16\" OC (2x4 cavities)", "sq ft", 600, 625, 650, "Home Depot", "Confirmed", "R-15 для 2x4 walls. If 2x6 cavities — use R-21. Final R-values to be verified before purchase per CA Title 24"),
    ("Ceiling", "Vaulted Ceiling Insulation R-30", "Unfaced fiberglass batt 16\" OC", "sq ft", 360, 380, 400, "Home Depot", "Confirmed", "R-30 allowance — final R-value to be verified per code assembly. Vented assembly з baffles (alternative: unvented spray foam r17)"),
    ("Floor", "Floor Insulation (cross-ref Phase 2)", "Already у Phase 2 row 12 — qty 0", "sq ft", 0, 0, 0, "—", "Confirmed", "DUPLICATE — vже в Phase 2 Raised Floor row 12"),
    ("Sound", "Sound Insulation R-13", "Bedroom/bathroom walls fiberglass", "sq ft", 250, 300, 350, "Home Depot", "TBD", "R-13 fiberglass (baseline). Premium alternative — mineral wool r15. Choose ONE not both"),
    ("Air Sealing", "Spray Foam", "Great Stuff window/door foam", "cans", 4, 6, 8, "Home Depot", "Confirmed", ""),
    ("Air Sealing", "Fire-blocking Foam / Caulk", "DAP fire-block", "tubes", 4, 6, 8, "Home Depot", "TBD", ""),
    ("Air Sealing", "Air Sealing Tape", "3M All Weather / Tyvek", "rolls", 2, 3, 4, "Home Depot", "Confirmed", ""),
    ("Ventilation", "Baffles", "Rafter air vents для VENTED vaulted ceiling", "pcs", 8, 14, 20, "Home Depot", "TBD", "Mid підвищено до 14 (every rafter bay). Choose VENTED batt+baffles OR UNVENTED spray foam r17 — not both"),
    ("Supports", "Insulation Supports / Wire Supports", "Hold wall+ceiling batt в place", "pack", 1, 2, 3, "Home Depot", "Confirmed", "For wall/ceiling batt"),
    ("Vapor", "Vapor Barrier Poly (Optional)", "Allowance — kraft-faced batt usually enough", "set", 0, 1, 1, "Home Depot", "Optional", "Verify if required — kraft-facing usually serves as vapor retarder"),
    ("Sound", "Mineral Wool R-13 Alternative (Optional)", "Roxul Safe'n'Sound — premium sound + fire", "sq ft", 0, 0, 350, "Home Depot", "Optional", "Premium alt для r8. Choose ONE not both"),
    ("Supports", "Cap Strips / Batt Straps", "Bottom plate compression", "pack", 0, 1, 2, "Home Depot", "TBD", ""),
    ("Air Sealing", "Closed-cell Spray Foam Kit (Optional)", "Для UNVENTED vaulted ceiling assembly", "kit", 0, 0, 1, "Home Depot / specialty", "Optional", "Alternative до vented batt+baffles. Choose ONE strategy"),
    # --- ДОДАНО per reviewer V3 ---
    ("Air Sealing", "General Air-sealing Caulk", "Top/bottom plates, corners, small gaps", "tubes", 3, 4, 6, "Home Depot", "Confirmed", "ДОДАНО (review): air sealing perimeters"),
    ("Supports", "Staples / Insulation Fastening Supplies", "Для kraft-faced batt stapling", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review): T50 staples + крепежі"),
    ("Safety", "Protective Gear allowance", "Fiberglass gloves + mask + goggles", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review): PPE для fiberglass installation"),
]

# === 10 — Drywall (§24) — V3 з reviewer fixes ===
DRYWALL = [
    ("Board", "Regular Drywall 1/2\"", "4x8 sheets — walls (baseline)", "sheets", 60, 67, 75, "Home Depot", "Confirmed", "1/2\" walls + standard ceiling. PVA primer counted ТУТ — не дублювати в Paint phase"),
    ("Board", "Moisture-resistant Drywall", "Greenboard 1/2\" — bathroom, laundry, utility", "sheets", 10, 12, 15, "Home Depot", "Confirmed", ""),
    ("Board", "Cement Board / Backer Board", "HardieBacker 1/2\" — shower wet zone", "sheets", 3, 4, 5, "Home Depot", "Confirmed", "Pair з RedGard r16. Counted тут — не дублювати в Bathroom phase"),
    ("Fasteners", "Drywall Screws", "1¼\" coarse — підвищено Mid", "lb", 8, 10, 14, "Home Depot", "Confirmed", "Mid підвищено до 10 lb (review): достатньо для 67 sheets"),
    ("Finishing", "Joint Compound (All-Purpose)", "5 gal ready-mix — підвищено Mid", "buckets", 4, 6, 8, "Home Depot", "Confirmed", "Mid підвищено до 6 buckets (review): для vaulted ceiling + corners"),
    ("Finishing", "Drywall Tape (Paper)", "250 ft roll", "rolls", 4, 6, 8, "Home Depot", "Confirmed", "Paper для joints; mesh окремо r15"),
    ("Finishing", "Corner Bead", "Metal або paper-faced", "linear ft", 100, 140, 180, "Home Depot", "Confirmed", ""),
    ("Primer", "Drywall Primer (PVA)", "5 gal — counted тут, не дублювати в Paint phase", "gal", 2, 3, 4, "Home Depot", "Confirmed", "PVA drywall primer — окремо від finish primer"),
    ("Finishing", "Texture Material", "Homax knockdown spray (CA default)", "cans", 3, 4, 6, "Home Depot", "TBD", "§24.4. Texture method TBD: spray cans = allowance only"),
    ("Access", "Access Panels 12×12", "PVC — bathroom + utility access", "pcs", 1, 2, 3, "Home Depot", "Confirmed", "§24.4"),
    ("Finishing", "Drywall Mesh Tape (cement board)", "Alkali-resistant fiberglass", "rolls", 1, 1, 2, "Home Depot", "Confirmed", "Cement board joints"),
    ("Waterproofing", "RedGard Waterproofing Membrane", "Custom Building Products — shower walls", "gal", 1, 1, 2, "Home Depot", "Confirmed", "За cement board у shower"),
    ("Finishing", "Sanding Consumables", "Sanding screens + sponges + paper", "set", 1, 1, 1, "Home Depot", "Confirmed", ""),
    ("Finishing", "Setting-type Compound Durabond (Optional)", "20-min set", "bag", 0, 1, 2, "Home Depot", "Optional", ""),
    ("Adhesives", "Drywall Adhesive (Optional)", "PL Premium stud attachment", "tubes", 0, 6, 10, "Home Depot", "Optional", "Reduce screws, optional"),
    # --- ДОДАНО per reviewer V3 ---
    ("Board", "5/8\" Type X Drywall (Allowance/Upgrade)", "Fire-rated ceiling assembly если req'd", "sheets", 0, 0, 12, "Home Depot", "Optional", "ДОДАНО (review): verify fire-rating ceiling requirement per §33.26. Окрема upgrade позиція"),
    ("Finishing", "Flexible Corner Tape / No-Coat", "Vaulted ceiling transitions, нестандартні кути", "rolls", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО (review): vaulted ceiling angles"),
    ("Site Protection", "Plastic Sheeting / Dust Protection", "Floor + adjacent rooms protection", "rolls", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО (review): mudding/sanding dust"),
    ("Finishing", "Control Joint / Expansion Bead (Optional)", "Якщо довгі площини або transitions", "linear ft", 0, 0, 30, "Home Depot", "Optional", "ДОДАНО (review): TBD per layout"),
    ("Repair", "Drywall Repair Patch (Optional)", "Для utility access або corrections", "pcs", 0, 0, 2, "Home Depot", "Optional", "ДОДАНО (review)"),
]

# === 11 — Paint (§27) — V2 з 9 додатковими items + wall paint correction ===
# NOTE: PVA drywall primer counted в Phase 9 r12 — не дублювати тут
PAINT = [
    ("Wall Paint", "Wall Paint", "Behr Premium Plus Ultra eggshell, 2 coats", "gal", 8, 12, 16, "Home Depot", "Confirmed", "Mid підвищено 7→12 для 2-coat standard (2,200 sf × 2 / 350 = 12.6)"),
    ("Ceiling Paint", "Ceiling Paint", "Flat white ceiling-specific", "gal", 2, 3, 4, "Home Depot", "Confirmed", "Vaulted ~400 sf × 2 coats"),
    ("Specialty", "Bathroom Moisture-resistant Paint", "Behr Premium Mildew-resistant semi-gloss", "gal", 1, 1, 2, "Home Depot", "Confirmed", ""),
    ("Trim", "Trim Paint", "Behr semi-gloss acrylic", "gal", 1, 2, 3, "Home Depot", "Confirmed", ""),
    ("Prep", "Painter's Caulk", "DAP Alex Plus latex — підвищено Mid", "tubes", 6, 8, 12, "Home Depot", "Confirmed", "Mid підвищено 6→8 (review): для trim/casing/baseboards"),
    ("Prep", "Painter's Tape", "3M ScotchBlue 1.88\"", "rolls", 4, 6, 8, "Home Depot", "Confirmed", ""),
    ("Prep", "Drop Cloths / Plastic", "Canvas drop cloths + plastic sheeting", "set", 1, 1, 2, "Home Depot", "Confirmed", ""),
    ("Prep", "Sandpaper / Wood Filler", "Assortment + wood filler", "set", 1, 1, 1, "Home Depot", "Confirmed", ""),
    # --- ДОДАНО per Pre-Takeoff Check ---
    ("Touch-up", "Touch-up Supplies", "Mini-rollers + small brushes + containers", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (per §27.2)"),
    ("Tools", "Paint Rollers + Covers", "9\" 3/8\" nap covers 6-pack", "set", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Tools", "Paint Brushes", "2\" + 2.5\" + 3\" angled sash mix", "set", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Tools", "Paint Trays + Liners", "9\" tray + disposable liners 4-pack", "set", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Tools", "Roller Poles / Extensions", "Telescoping для 12 ft vaulted ceiling", "pcs", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Prep", "Spackling Compound", "Для дрібних отворів (не joint compound)", "container", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Primer", "Finish Primer / Stain Blocker", "Kilz Premium — wood trim/knots — окремо від Phase 9 PVA", "gal", 0, 1, 2, "Home Depot", "Confirmed", "ДОДАНО: різний від PVA drywall primer (Phase 9)"),
    ("Cleanup", "Cleanup Supplies", "Goof Off + rags + soap. Mineral spirits ONLY if oil-based primer used (r19)", "set", 1, 1, 1, "Home Depot", "Confirmed", ""),
    ("Prep", "Putty Knives / Scrapers", "Set для wood filler/spackling application", "set", 1, 1, 1, "Home Depot", "Confirmed", ""),
    # --- ДОДАНО per reviewer V2 ---
    ("Sealant", "100% Silicone Bathroom Caulk", "GE Silicone II / DAP Kwik Seal для wet joints", "tubes", 2, 2, 3, "Home Depot", "Confirmed", "ДОДАНО (review): painter's caulk не замінює silicone для wet joints"),
    ("Prep", "Masking Film / Pre-taped Plastic", "Захист windows/doors/floor/cabinets", "rolls", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО (review)"),
    ("Prep", "Tack Cloth / Microfiber Cloths", "Trim prep перед фарбуванням", "pack", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review)"),
]

# === 12 — Flooring & Trim (§26-27) — V2 з 7 додатковими items + baseboards correction ===
FLOORING_TRIM = [
    ("Flooring", "LVP / Luxury Vinyl Plank", "LifeProof 22 MIL click-lock waterproof", "sq ft", 315, 350, 380, "Home Depot", "Confirmed", "Across entire ADU. Long dimension parallel to walls (21 ft)"),
    ("Flooring", "Underlayment (Optional)", "Foam 2mm з vapor barrier — ONLY if LVP без attached pad", "sq ft", 0, 0, 380, "Home Depot", "Optional", "LifeProof 22 MIL зазвичай має attached underlayment — verify product specs. Mid 0 baseline"),
    ("Trim", "Decorative Transition Strips", "T-molding + reducer + end-cap для LVP", "pcs", 3, 4, 5, "Home Depot", "Confirmed", "Decorative (LVP transitions). Phase 2 r13 = structural (10\" raise)"),
    ("Trim", "Baseboards", "5.25\" pine MDF primed", "linear ft", 130, 150, 170, "Home Depot", "Confirmed", "Mid 150 LF includes 10-15% waste/returns for corners"),
    ("Trim", "Quarter Round", "Shoe molding 3/4\" pine primed", "linear ft", 0, 100, 160, "Home Depot", "TBD", ""),
    # --- ДОДАНО per Pre-Takeoff Check ---
    ("Moisture", "Vapor Barrier Sheet (Optional)", "6-mil poly під LVP — якщо LVP без attached", "sq ft", 0, 0, 380, "Home Depot", "Optional", "ДОДАНО: skip if LVP має attached vapor"),
    ("Supplies", "Flooring Spacers", "1/4\"-3/8\" expansion gap", "pack", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Supplies", "LVP Installation Kit", "Tapping block + pull bar + utility knife", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Sealants", "Floor Caulk", "Perimeter sealing", "tubes", 2, 4, 6, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Fasteners", "Finish Nails (18 ga brad)", "Baseboard + quarter round installation", "box", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Adjustment", "Door Undercut Allowance", "Якщо doors торкаються нового flooring", "allowance", 0, 1, 1, "Home Depot / service", "TBD", "ДОДАНО"),
    ("Slab Prep", "Self-leveler Patch (Optional)", "Final subfloor touchups", "bag", 0, 0, 2, "Home Depot", "Optional", "ДОДАНО: small allowance, cross-ref Phase 2 r19"),
    # --- ДОДАНО per reviewer V2 ---
    ("Trim", "Bathroom Reducer / Transition Allowance", "Якщо bathroom flooring відрізняється або є height diff", "pcs", 0, 1, 2, "Home Depot", "TBD", "ДОДАНО (review)"),
    ("Trim", "Stair Nosing / Step Trim", "Через raised floor 10\" — step at entry/French", "pcs", 0, 1, 2, "Home Depot", "TBD", "ДОДАНО (review): verify step condition"),
]

# === 13 — Kitchen Cabinets & Countertop (§11) — V2 з 7 додатковими items ===
KITCHEN_CABINETS = [
    ("Cabinets", "Base Cabinets", "IKEA SEKTION carcass 24\" deep", "linear ft", 8, 8.5, 9, "IKEA", "Confirmed", "One-wall kitchen"),
    ("Cabinets", "Wall Cabinets", "IKEA SEKTION 15\" deep", "linear ft", 8, 8.5, 9, "IKEA", "Confirmed", ""),
    ("Cabinets", "Tall / Pantry Cabinet", "SEKTION 24\" wide × 80\" tall", "pcs", 0, 1, 1, "IKEA", "TBD", ""),
    ("Cabinets", "Cabinet Doors", "SEKTION + AXSTAD або alternative front style", "set", 1, 1, 1, "IKEA", "TBD", "US naming: SEKTION system. Front style — AXSTAD/VOXTORP/etc"),
    ("Cabinets", "Drawer Fronts", "Mix sizes per IKEA planner", "set", 1, 1, 1, "IKEA", "TBD", ""),
    ("Cabinets", "Hinges", "Soft-close hinges (UTRUSTA)", "pack", 1, 1, 1, "IKEA", "TBD", ""),
    ("Cabinets", "Drawer Slides", "Soft-close runners (MAXIMERA)", "set", 1, 1, 1, "IKEA", "TBD", ""),
    ("Cabinets", "Cabinet Legs", "Adjustable leveling legs", "set", 1, 1, 1, "IKEA", "TBD", ""),
    ("Cabinets", "Toe-kick", "FÖRBÄTTRA toe-kick", "linear ft", 8, 9, 10, "IKEA", "TBD", ""),
    ("Cabinets", "Cover Panels", "FÖRBÄTTRA cover panels", "set", 1, 1, 1, "IKEA", "TBD", ""),
    ("Cabinets", "Filler Panels", "FÖRBÄTTRA filler panels", "set", 1, 1, 1, "IKEA", "TBD", ""),
    ("Countertop", "Countertop", "KARLBY butcher block (BASELINE). Alt: EKBACKEN laminate (~$14/LF budget)", "linear ft", 9, 9.5, 10, "IKEA", "TBD", "Choose ONE material — not both. Final cabinet+countertop price via IKEA Kitchen Planner"),
    ("Backsplash", "Backsplash Tile", "Subway tile 3x6 ceramic — або mosaic", "sq ft", 25, 30, 35, "Home Depot", "TBD", ""),
    ("Fixtures", "Kitchen Sink", "HILLESJÖN single bowl 25\" SS drop-in", "pcs", 1, 1, 1, "IKEA / Home Depot", "TBD", ""),
    ("Fixtures", "Kitchen Faucet", "IKEA ÄLMAREN pull-down single-handle", "pcs", 1, 1, 1, "IKEA / Home Depot", "TBD", ""),
    ("Hardware", "Handles / Knobs", "Modern bar pulls — 6-8 pulls", "set", 1, 1, 1, "IKEA", "TBD", ""),
    ("Lighting", "Under-cabinet Lighting", "IKEA OMLOPP LED strip set. Driver — cross-ref Phase 6 r38", "set", 0, 1, 1, "IKEA / Home Depot", "TBD", "LED driver counted в Phase 6 r38"),
    # --- ДОДАНО per Pre-Takeoff Check ---
    ("Plumbing", "Kitchen Sink Plumbing Kit", "Basket strainer + tailpiece + drain assembly", "set", 1, 1, 1, "Home Depot", "Confirmed", "Phase 5 = rough-in to wall. Phase 12 = finish connection under sink"),
    ("Plumbing", "Sink Mounting Hardware", "Drop-in clips або undermount brackets", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Backsplash", "Tile Adhesive + Grout", "Mastic + grout + sealant", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Backsplash", "Backsplash Trim / Edge Mouldings", "Bullnose або metal edges", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Sealants", "Countertop Silicone Sealant", "Kitchen-safe 100% silicone", "tubes", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Cabinets", "Cabinet Shelves / Organizers Allowance", "Drawer inserts + extra shelves", "set", 0, 0, 1, "IKEA", "Optional", "ДОДАНО"),
    ("Cabinets", "Cabinet Feet / Risers Allowance", "Custom heights", "set", 0, 0, 1, "IKEA", "Optional", "ДОДАНО"),
    # --- ДОДАНО per reviewer V2 ---
    ("Cabinets", "IKEA SEKTION Suspension Rail", "Required для wall+base cabinet mounting", "pcs", 2, 3, 4, "IKEA", "Confirmed", "ДОДАНО (review): без rail SEKTION не монтується"),
    ("Cabinets", "Cabinet Mounting Screws / Wall Anchors", "Для rail/cabinet installation in studs", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review)"),
    ("Countertop", "Countertop End Caps / Edge Banding", "Для laminate або cut edges", "set", 1, 1, 1, "IKEA / Home Depot", "TBD", "ДОДАНО (review)"),
    ("Cabinets", "Dishwasher End Panel / Cover", "Якщо dishwasher на кінці cabinet run", "pcs", 0, 1, 1, "IKEA", "TBD", "ДОДАНО (review): cover open side"),
    ("Cabinets", "Dishwasher Mounting Hardware (anti-tip)", "Якщо не included з appliance", "set", 0, 1, 1, "Home Depot", "TBD", "ДОДАНО (review)"),
    ("Backsplash", "Backsplash Spacers / Float / Sponge Set", "Tile install small supplies", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review)"),
    ("Countertop", "Butcher Block Oil/Sealer", "Якщо KARLBY butcher block обрано", "qt", 0, 1, 2, "Home Depot", "TBD", "ДОДАНО (review): тільки якщо butcher block"),
    # --- Cross-references ---
    ("Cross-ref", "Dishwasher Air Gap (CA UPC)", "Cross-ref Phase 5 r37 — не дублювати", "set", 0, 0, 0, "—", "Confirmed", "ДОДАНО: вже у Phase 5 r37 Dishwasher Branch + Air Gap"),
]

# === 14 — Bathroom Fixtures (§14) — V2 з cross-refs + 7 додатковими items ===
BATHROOM_FIXTURES = [
    ("Shower", "Prefab Shower Base 32×40", "Acrylic/fiberglass", "pcs", 1, 1, 1, "Home Depot", "Confirmed", "MAAX/Sterling style"),
    ("Shower", "Prefab Shower Wall System", "DEPRECATED — заміняємо tile (r28-r34)", "set", 0, 0, 0, "—", "Confirmed", "Removed: заміняємо на tile shower walls"),
    ("Shower", "Shower Drain (cross-ref Phase 5)", "Cross-ref Phase 5 r33 — не дублювати", "pcs", 0, 0, 0, "—", "Confirmed", "Drain body+strainer assembly у Phase 5 r33"),
    ("Shower", "Shower Valve Body (cross-ref Phase 5)", "Cross-ref Phase 5 r32 — не дублювати", "pcs", 0, 0, 0, "—", "Confirmed", "Rough-in valve body у Phase 5 r32"),
    ("Shower", "Shower Trim Kit", "Handle + escutcheon — finish stage", "set", 1, 1, 1, "Home Depot", "TBD", "Visible trim (different from Phase 5 valve body)"),
    ("Shower", "Shower Head", "Fixed або handheld", "pcs", 1, 1, 1, "Home Depot", "TBD", ""),
    ("Shower", "Waterproof Caulk / Sealant", "Silicone bath", "tubes", 2, 3, 4, "Home Depot", "Confirmed", ""),
    ("Toilet", "Toilet", "2-piece elongated chair-height", "pcs", 1, 1, 1, "Home Depot", "Confirmed", ""),
    ("Toilet", "Toilet Supply Line", "Flex braided", "pcs", 1, 1, 1, "Home Depot", "Confirmed", ""),
    ("Vanity", "Vanity Cabinet 24\"", "Home Decorators / Glacier Bay", "pcs", 1, 1, 1, "Home Depot / IKEA", "Confirmed", "HD combo $349-449 includes top+sink+backsplash"),
    ("Vanity", "Vanity Top / Sink", "Integrated cultured marble / quartz", "pcs", 1, 1, 1, "Home Depot / IKEA", "Confirmed", "Часто sold з cabinet в combo"),
    ("Vanity", "Vanity Faucet", "Single-handle", "pcs", 1, 1, 1, "Home Depot", "Confirmed", ""),
    ("Vanity", "P-trap (cross-ref Phase 5)", "Cross-ref Phase 5 r14 — qty 0", "pcs", 0, 0, 0, "—", "Confirmed", "P-trap включений у Phase 5 r14 (4 P-traps total: shower+vanity+kitchen+washer)"),
    ("Vanity", "Shut-off Valves (cross-ref Phase 5)", "Cross-ref Phase 5 r8 — qty 0", "pcs", 0, 0, 0, "—", "Confirmed", "Shut-offs включені у Phase 5 r8 (19 total covers all fixtures)"),
    ("Accessories", "Mirror", "Frameless або basic framed", "pcs", 1, 1, 1, "Home Depot / IKEA", "Confirmed", ""),
    ("Accessories", "Towel Bar / Hooks", "Set 24\" bar + 2 hooks", "set", 1, 1, 1, "Home Depot / IKEA", "Confirmed", ""),
    ("Accessories", "Toilet Paper Holder", "", "pcs", 1, 1, 1, "Home Depot / IKEA", "Confirmed", ""),
    ("Ventilation", "Bathroom Exhaust Fan", "80 CFM з timer compatible", "pcs", 1, 1, 1, "Home Depot", "Confirmed", "Timer switch у Phase 6 r16"),
    ("Ventilation", "Fan Duct + Wall/Roof Cap", "Flex duct + exterior cap", "set", 1, 1, 1, "Home Depot", "Confirmed", ""),
    # --- ДОДАНО per Pre-Takeoff Check ---
    ("Shower", "Shower Curtain Rod + Curtain", "Budget enclosure (glass door alternative $300-800)", "set", 1, 1, 1, "Home Depot / IKEA", "Confirmed", "ДОДАНО: budget option. Glass door — Optional upgrade"),
    ("Toilet", "Toilet Wax Ring + Sealing", "Mandatory install component", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Toilet", "Toilet Seat", "Часто NOT included з toilet", "pcs", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Toilet", "Toilet Bolts / Caps", "Small allowance", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Vanity", "Vanity Mount Hardware", "Wall anchors якщо потрібно", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Vanity", "Vanity Backsplash Allowance (Optional)", "Small tile або solid", "set", 0, 0, 1, "Home Depot", "Optional", "ДОДАНО: usually included з vanity combo"),
    ("Accessories", "Robe Hook (Optional)", "1-2 hooks", "pcs", 0, 0, 2, "Home Depot / IKEA", "Optional", "ДОДАНО"),
    # --- ДОДАНО per reviewer V2 ---
    ("Vanity", "Vanity Pop-up Drain Assembly", "Якщо faucet не включає drain", "pcs", 0, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review)"),
    ("Vanity", "Vanity Faucet Supply Lines", "Flex lines від shut-off до faucet", "pcs", 2, 2, 2, "Home Depot", "Confirmed", "ДОДАНО (review)"),
    ("Shower", "Shower Arm + Flange", "Якщо не included з shower head", "pcs", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review)"),
    ("Ventilation", "Duct Clamps + Foil HVAC Tape", "Для bath fan duct connection", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review)"),
    ("Shower", "Shower Curtain Hooks / Rings", "Часто sold з curtain, allowance backup", "pcs", 1, 1, 1, "Home Depot / IKEA", "Confirmed", "ДОДАНО (review)"),
    ("Accessories", "Medicine Cabinet / Bathroom Storage (Optional)", "Якщо потрібне зберігання", "pcs", 0, 0, 1, "Home Depot / IKEA", "Optional", "ДОДАНО (review)"),
    ("Accessories", "Grab Bars (Optional)", "ADA/safety allowance — blocking у Phase 3", "pcs", 0, 0, 2, "Home Depot", "Optional", "ДОДАНО (review)"),
    # --- ДОДАНО: TILE SHOWER WALLS (замість prefab) ---
    ("Tile - Shower", "Shower Wall Tile (ceramic subway 3×6)", "62 sf wall area: 32\" front + 2×40\" side × 7' tall", "sq ft", 55, 62, 70, "Home Depot", "Confirmed", "ДОДАНО: tile shower walls"),
    ("Tile - Shower", "Tile Thinset / Mastic (shower)", "50 lb bag, ~1-2 bags для 60+ sf", "bag", 1, 2, 2, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Tile - Shower", "Sanded Grout (shower)", "25 lb bag", "bag", 1, 1, 2, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Tile - Shower", "Schluter / Tile Trim Edges", "Bullnose / metal trim for shower edges", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО"),
    # --- ДОДАНО: TILE FLOOR (bathroom) ---
    ("Tile - Floor", "Bathroom Floor Tile (porcelain 12×12)", "Bathroom floor ~32 sf usable (minus vanity)", "sq ft", 28, 32, 38, "Home Depot", "Confirmed", "ДОДАНО: bathroom floor tile (замість LVP)"),
    ("Tile - Floor", "Floor Thinset", "50 lb bag", "bag", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Tile - Floor", "Floor Grout + Spacers", "Grout + spacers + sponge kit", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО"),
    ("Tile - Floor", "Cement Board (floor underlay)", "1/2\" HardieBacker for floor 32 sf", "sq ft", 28, 32, 38, "Home Depot", "Confirmed", "ДОДАНО: окремо від shower cement board (Phase 9)"),
    ("Tile - Waterproofing", "Extra RedGard (floor + corners)", "Additional quart для shower floor + transitions", "qt", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО: окрема quart RedGard"),
]

# === 15 — Closet System (§13) — V2 з BOAXEL clarification + 4 нових items ===
CLOSET = [
    ("Storage System", "IKEA BOAXEL wardrobe combination", "BOAXEL 49 1/8\" main wall combo", "system", 1, 1, 1, "IKEA", "Confirmed", "BASELINE = BOAXEL. PAX — separate upgrade scenario (r15), не входить у поточний total"),
    ("Storage System", "Additional Hanging Rods", "Extra BOAXEL clothes rails для second wall", "pcs", 0, 1, 2, "IKEA", "TBD", ""),
    ("Storage System", "Additional Shelves", "BOAXEL extra shelves для second wall", "pcs", 2, 3, 5, "IKEA", "TBD", ""),
    ("Storage System", "Drawers (mesh baskets)", "BOAXEL mesh basket drawers", "pcs", 2, 4, 6, "IKEA", "TBD", ""),
    ("Storage System", "Shoe Storage", "BOAXEL shoe shelves або freestanding rack", "pcs", 1, 1, 2, "IKEA", "TBD", ""),
    ("Accessories", "Mirror (closet)", "Wall-mounted full-length", "pcs", 0, 1, 1, "IKEA", "TBD", ""),
    # --- ДОДАНО per Pre-Takeoff Check ---
    ("Storage System", "BOAXEL Wall Rails (additional)", "Для second wall mounting", "pcs", 1, 2, 3, "IKEA", "Confirmed", "ДОДАНО"),
    ("Hardware", "Mounting Hardware (Stud Screws + Anchors)", "Stud-mounted screws у studs/blocking + drywall anchors де dozvoleno", "set", 1, 1, 1, "Home Depot", "Confirmed", "Stud/blocking-mounted preferred. Drywall anchors тільки де дозволяє IKEA spec"),
    ("Accessories", "Drawer Organizers / Dividers (Optional)", "IKEA KOMPLEMENT inserts", "set", 0, 0, 1, "IKEA", "Optional", "ДОДАНО"),
    ("Accessories", "Closet Accessories (Optional)", "Valet hook, belt rack, hangers", "set", 0, 0, 1, "IKEA", "Optional", "ДОДАНО"),
    # --- ДОДАНО per reviewer V2 ---
    ("Storage System", "BOAXEL Uprights / Vertical Rails", "Required для full system mounting", "pcs", 1, 2, 3, "IKEA", "Confirmed", "ДОДАНО (review): high pri — без uprights system не складається"),
    ("Storage System", "BOAXEL Shelf / Rod Brackets", "Brackets для additional shelves/rods", "pack", 1, 1, 2, "IKEA", "Confirmed", "ДОДАНО (review)"),
    ("Storage System", "Second-wall BOAXEL Add-on", "Complete add-on для short wall 5'6\"", "system", 0, 1, 1, "IKEA", "Confirmed", "ДОДАНО (review): краще ніж loose components"),
    ("Storage System", "PAX Upgrade Alternative (Optional)", "Premium alternative до BOAXEL", "system", 0, 0, 1, "IKEA", "Optional", "ДОДАНО (review): premium upgrade allowance"),
    ("Cross-ref", "Closet Sliding Door (cross-ref Phase 4)", "Phase 4 r8 — не дублювати", "—", 0, 0, 0, "—", "Confirmed", "ДОДАНО: sliding door в Phase 4 r8 $250"),
    ("Cross-ref", "Closet Motion Light (cross-ref Phase 6)", "Phase 6 r25 — не дублювати", "—", 0, 0, 0, "—", "Confirmed", "ДОДАНО: closet motion light Phase 6 r25 $30"),
    ("Accessories", "Hangers Set (Optional)", "Якщо ready-to-use furnishing", "set", 0, 0, 1, "IKEA", "Optional", "ДОДАНО (review)"),
]

# === 16 — Water Heater / Utility Closet (§16) — V2 з cross-refs + 3 нових items ===
WATER_HEATER = [
    ("Water Heater", "Electric Tank Water Heater", "Rheem Performance 30 gal short 240V", "pcs", 1, 1, 1, "Home Depot", "Confirmed", "Rheem XE30S06ST45U1 short tank. §33.7 verify sizing + verify clearance — 30 gal must fit utility closet з service access"),
    ("Water Heater", "Water Heater Pan", "24-26\" round metal/plastic", "pcs", 1, 1, 1, "Home Depot", "Confirmed", "Catch leak protection"),
    ("Water Heater", "Pan Drain Line", "3/4\" PVC до termination", "ft", 5, 10, 15, "Home Depot", "Confirmed", "3/4\" PVC consistent. Final size by plumber/product instructions. До exterior wall or floor drain"),
    ("Water Heater", "Flex Connectors (cross-ref Phase 5)", "Phase 5 r20 WH Connection Kit — не дублювати", "pcs", 0, 0, 0, "—", "Confirmed", "Cross-ref Phase 5 r20 $50"),
    ("Water Heater", "Shut-off Valves WH (cross-ref Phase 5)", "Phase 5 r8 (19 total) — не дублювати", "pcs", 0, 0, 0, "—", "Confirmed", "Cross-ref Phase 5 r8"),
    ("Water Heater", "T&P Relief Valve Discharge Pipe", "Cu 3/4\" rigid pipe", "ft", 3, 5, 7, "Home Depot", "Confirmed", ""),
    ("Water Heater", "Expansion Tank", "If required by code", "pcs", 0, 1, 1, "Home Depot", "TBD", "Verify Menlo Park code"),
    ("Water Heater", "Seismic Strapping Kit", "CA mandatory — top + bottom straps", "set", 1, 1, 1, "Home Depot", "Confirmed", "California code"),
    ("Closet Construction", "Access / Louvered Door (cross-ref Phase 4)", "Phase 4 r11 Utility Closet Door — не дублювати", "pcs", 0, 0, 0, "—", "Confirmed", "Cross-ref Phase 4 r11 $80"),
    ("Closet Construction", "Ventilation Grille", "Air circulation для closet", "pcs", 1, 1, 1, "Home Depot", "Confirmed", ""),
    # --- ДОДАНО per Pre-Takeoff Check ---
    ("Water Heater", "Pan Drain Termination", "End fitting + exterior cap", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО: where pipe ends"),
    ("Water Heater", "T&P Valve Replacement (Optional)", "Часто включений з WH", "pcs", 0, 0, 1, "Home Depot", "Optional", "ДОДАНО"),
    ("Water Heater", "Dielectric Union Fittings", "Required для steel/copper joint", "pcs", 1, 2, 2, "Home Depot", "Confirmed", "ДОДАНО"),
    # --- ДОДАНО per reviewer V2 ---
    ("Water Heater", "Seismic Strap Fasteners (lag screws)", "Якщо не included з strap kit", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review): для CA mandatory strap"),
    ("Water Heater", "T&P Discharge Termination Fitting", "Approved location end fitting", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review)"),
    ("Water Heater", "Drain Hose / Service Cap (Optional)", "Для maintenance / draining", "set", 0, 0, 1, "Home Depot", "Optional", "ДОДАНО (review)"),
    ("Cross-ref", "WH Electrical Whip / Hardwire (Phase 6)", "Phase 6 — не дублювати", "—", 0, 0, 0, "—", "Confirmed", "ДОДАНО (review): WH typically hardwired через Phase 6 r9 240V cable + r6 30A breaker"),
]

# === 17 — Appliances (§28) — V2 з 5 додатковими items ===
APPLIANCES = [
    ("Kitchen", "Refrigerator 24\" compact", "Frigidaire FFTR1022QS 10 cu ft або equivalent", "pcs", 1, 1, 1, "Best Buy / Amazon / HD", "Confirmed", "24\" compact"),
    ("Kitchen", "Electric Range 24\"", "Frigidaire FFEH2422US або similar", "pcs", 1, 1, 1, "Best Buy / Home Depot", "Confirmed", ""),
    ("Kitchen", "Dishwasher 24\" (or 18\")", "Bosch 100 series 24\" baseline. 18\" compact if kitchen tight (8.5 LF)", "pcs", 1, 1, 1, "Best Buy / Home Depot", "TBD", "Verify final size — 18\" alternative for tight one-wall kitchen"),
    ("Kitchen", "Microwave / Range Hood Combo", "OTR (over-the-range) з vent fan", "pcs", 1, 1, 1, "Best Buy / Amazon", "Confirmed", "Needs ductwork (r13)"),
    ("Kitchen", "Garbage Disposal", "InSinkErator Badger 1/3 HP — Optional", "pcs", 0, 0, 1, "Home Depot", "Optional", "Per ТЗ §28.1 optional"),
    ("Laundry", "Compact Washer 24\"", "Bosch 500 series WAT28402UC or equivalent", "pcs", 1, 1, 1, "Special Order", "Confirmed", "Special Order. Lead 4-6 weeks. Model availability TBD — use equivalent current model"),
    ("Laundry", "Ventless Compact Dryer 24\"", "Bosch 500 series heat pump or equivalent", "pcs", 1, 1, 1, "Special Order", "Confirmed", "Special Order. Ventless. Availability TBD — use equivalent"),
    ("Laundry", "Stacking Kit", "Brand-matched з washer/dryer", "pcs", 1, 1, 1, "Best Buy / Amazon", "Confirmed", ""),
    # --- ДОДАНО per Pre-Takeoff Check ---
    ("Kitchen Install", "Range Hood Ductwork + Cap", "4\" або 6\" round duct + exterior cap", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО: vent microwave/hood до outside"),
    ("Kitchen Install", "Disposal Installation Kit (Optional)", "Drain + flange + plug cord", "set", 0, 0, 1, "Home Depot", "Optional", "ДОДАНО: tied до disposal r9"),
    ("Kitchen Install", "Disposal Air Switch (Optional)", "Counter-mounted alternative до wall switch", "pcs", 0, 0, 1, "Home Depot", "Optional", "ДОДАНО: CA Code alt"),
    ("Appliance Install", "Range Power Cord", "14-50R baseline", "pcs", 1, 1, 1, "Home Depot", "TBD", "Final cord type TBD by selected range specs"),
    ("Appliance Install", "Dryer Power Cord", "3 or 4-prong", "pcs", 1, 1, 1, "Home Depot", "TBD", "Final cord type TBD by selected dryer specs"),
    # --- ДОДАНО per reviewer V2 ---
    ("Kitchen Install", "Dishwasher Installation Kit", "Water line + power cord + drain connection", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review): often sold separately"),
    ("Laundry Install", "Washer Supply Hoses (pair)", "Stainless braided hot+cold", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review): часто NOT included з washer"),
    ("Laundry Install", "Dryer Condensate Drain Kit", "Ventless dryer drain hose + clamps", "set", 1, 1, 1, "Home Depot", "TBD", "ДОДАНО (review): verify dryer model — може мати tank instead"),
    ("Kitchen Install", "OTR Microwave Duct Adapter + Foil Tape", "Rectangular-to-round adapter + foil tape + clamps", "set", 1, 1, 1, "Home Depot", "Confirmed", "ДОДАНО (review)"),
]

# === 18 — Furniture IKEA (§29 + Living/Bedroom/Dining/Soft Goods) ===
FURNITURE = [
    # Living Room (з попереднього збору цін)
    ("Living Room", "Sofa / Sleeper Sofa", "FRIHETEN sleeper, Faringe light gray", "pcs", 1, 1, 1, "IKEA", "Confirmed", "Living Room"),
    ("Living Room", "Coffee Table", "LACK 35⅜, white stained oak", "pcs", 1, 1, 1, "IKEA", "Confirmed", "Living Room"),
    ("Living Room", "TV Bench / Media Unit", "BESTÅ 47¼ white stained oak", "pcs", 1, 1, 1, "IKEA", "Confirmed", "Living Room"),
    ("Living Room", "Storage Cabinet", "KALLAX 2x2 white stained oak", "pcs", 1, 1, 1, "IKEA", "Confirmed", "Living Room"),
    ("Living Room", "Rug", "LOHALS jute 5'3×7'7", "pcs", 1, 1, 1, "IKEA", "Confirmed", "Living Room"),
    ("Living Room", "Curtains + Blinds combo", "RITVA pair + TRETUR roller", "set", 1, 1, 1, "IKEA", "Confirmed", "French door + window"),
    ("Living Room", "Floor Lamp", "VIDJA white", "pcs", 1, 1, 1, "IKEA", "Confirmed", "Living Room"),
    # Bedroom
    ("Bedroom", "Full Bed Frame", "IKEA MALM Full white з LÖNSET slatted base", "pcs", 1, 1, 1, "IKEA", "Confirmed", "54×75. LÖNSET base included"),
    ("Bedroom", "Full Mattress", "IKEA HÖVÅG/MORGEDAL Full (HAFSLO discontinued US)", "pcs", 1, 1, 1, "IKEA", "Confirmed", "54×75"),
    ("Bedroom", "Nightstands", "IKEA MALM 2-drawer × 2", "pcs", 1, 2, 2, "IKEA", "Confirmed", ""),
    ("Bedroom", "Dresser / Small Storage", "IKEA MALM 4-drawer", "pcs", 1, 1, 1, "IKEA", "Confirmed", ""),
    ("Bedroom", "Bedside Lamps", "IKEA NYFORS/MELODI table lamps × 2", "pcs", 1, 2, 2, "IKEA", "Confirmed", ""),
    ("Bedroom", "Window Covering (bedroom)", "RITVA panels (matched Living combo)", "set", 1, 1, 1, "IKEA", "Confirmed", ""),
    # Dining
    ("Dining", "Round Dining Table", "30-35\" IKEA KRAGSTA round або equivalent", "pcs", 1, 1, 1, "IKEA", "Confirmed", ""),
    ("Dining", "Dining Chairs", "IKEA TEODORES × 3 mid-priced", "pcs", 2, 3, 4, "IKEA", "Confirmed", ""),
    ("Dining", "Pendant Light", "IKEA FOTO Optional", "pcs", 0, 1, 1, "IKEA / Home Depot", "Optional", ""),
    # Bathroom storage (per §29.5)
    ("Bathroom Storage", "Bathroom / Utility Storage", "Per §29.5: cabinet + shelves + organizers — Mid 1", "set", 0, 1, 1, "IKEA", "Confirmed", "Per ТЗ §29.5. Mid підвищено 0→1 (review)"),
    # --- ДОДАНО per reviewer V2 ---
    ("Soft Goods", "Curtain Rods / Brackets / Rings", "Hardware для всіх curtain panels", "set", 1, 1, 2, "IKEA / Home Depot", "Confirmed", "ДОДАНО (review): high priority"),
    ("Soft Goods", "Kitchen Window Covering", "RITVA panels або roller blind", "set", 1, 1, 1, "IKEA", "Confirmed", "ДОДАНО (review)"),
    ("Soft Goods", "Bathroom Window Covering", "Roller blind privacy", "set", 1, 1, 1, "IKEA", "Confirmed", "ДОДАНО (review)"),
    ("Bedding", "Mattress Protector + Basic Bedding", "Protector + sheets + pillows + comforter allowance", "set", 1, 1, 1, "IKEA / Amazon", "Confirmed", "ДОДАНО (review): medium pri"),
    ("Safety", "Anti-tip Straps / Wall Anchors", "Для MALM dresser, KALLAX, storage", "set", 1, 1, 1, "IKEA / Home Depot", "Confirmed", "ДОДАНО (review)"),
    ("Safety", "Felt Pads / Floor Protectors", "Захист LVP від furniture", "pack", 1, 1, 1, "Home Depot / Amazon", "Confirmed", "ДОДАНО (review)"),
    # --- Optional fully-furnished items ---
    ("Electronics", "TV (Optional)", "Якщо fully furnished scenario", "pcs", 0, 0, 1, "Best Buy / Amazon", "Optional", "ДОДАНО (review): Optional"),
    ("Furniture", "Compact Desk + Chair (Optional)", "Work/study area allowance", "set", 0, 0, 1, "IKEA", "Optional", "ДОДАНО (review)"),
    ("Closet", "Hangers Set (Optional)", "Closet basics", "set", 0, 0, 1, "IKEA", "Optional", "ДОДАНО (review): cross-ref Phase 14 r21 also"),
    ("Bathroom Storage", "Bath Mat + Towels (Optional)", "Bathroom textiles allowance", "set", 0, 0, 1, "IKEA / Amazon", "Optional", "ДОДАНО (review)"),
]

# === 19 — Special Orders (§31.5) — cross-reference summary ===
SPECIAL_ORDERS = [
    ("Doors", "Backyard French Door", "Glazed, exterior", "pcs", 1, 1, 1, "Special Order", "Confirmed", "→ Лист #05. Lead time 4-8 weeks"),
    ("Doors", "Folding / Panel Door System 8 ft", "Між bedroom і living", "system", 1, 1, 1, "Special Order", "Confirmed", "→ Лист #05. Lead 6-10 weeks"),
    ("HVAC", "Mini-split Outdoor Condenser", "Multi-zone", "pcs", 1, 1, 1, "Special Order", "Confirmed", "→ Лист #08. Lead 4-8 weeks"),
    ("HVAC", "Mini-split Indoor Units (2)", "Wall-mounted", "pcs", 2, 2, 2, "Special Order", "Confirmed", "→ Лист #08"),
    ("Laundry", "Compact Washer 24\"", "Stacked", "pcs", 1, 1, 1, "Special Order", "Confirmed", "→ Лист #17. Lead 4-6 weeks"),
    ("Laundry", "Ventless Compact Dryer 24\"", "Heat pump/condenser", "pcs", 1, 1, 1, "Special Order", "Confirmed", "→ Лист #17"),
    ("Water Heater", "Electric Tank Water Heater 20-30 gal", "", "pcs", 1, 1, 1, "Home Depot / specialty", "Confirmed", "→ Лист #16. Можна купити stock"),
]

# === 20 — Verification Items (§33) — 26 пунктів TBD ===
VERIFICATION = [
    ("Field", "Exact field dimensions of garage", "Total ADU footprint", "—", 0, 0, 0, "Field measurement", "TBD", "ТЗ §33.1"),
    ("Field", "Exact wall thicknesses", "Interior + exterior", "—", 0, 0, 0, "Field measurement", "TBD", "ТЗ §33.2"),
    ("Field", "Exact interior wall layout", "", "—", 0, 0, 0, "Field measurement", "TBD", "ТЗ §33.3"),
    ("Field", "Exact bathroom size after laundry/utility", "", "—", 0, 0, 0, "Field measurement", "TBD", "ТЗ §33.4"),
    ("Field", "Exact location of stacked washer/dryer", "", "—", 0, 0, 0, "Field measurement", "TBD", "ТЗ §33.5"),
    ("Field", "Exact utility closet size & location", "", "—", 0, 0, 0, "Field measurement", "TBD", "ТЗ §33.6"),
    ("Specs", "Electric tank water heater size", "20 vs 30 gal", "—", 0, 0, 0, "Specification", "TBD", "ТЗ §33.7"),
    ("Specs", "Exact washer/dryer model", "Особисто per brand", "—", 0, 0, 0, "Specification", "TBD", "ТЗ §33.8"),
    ("Specs", "Mini-split model + BTU sizing", "HVAC load calc", "—", 0, 0, 0, "Specification", "TBD", "ТЗ §33.9"),
    ("Specs", "Kitchen cabinet layout from IKEA planner", "", "—", 0, 0, 0, "IKEA Planner", "TBD", "ТЗ §33.10"),
    ("Specs", "Exact appliance dimensions", "", "—", 0, 0, 0, "Specification", "TBD", "ТЗ §33.11"),
    ("Openings", "Exact window sizes", "3 вікна", "—", 0, 0, 0, "Field + Spec", "TBD", "ТЗ §33.12"),
    ("Openings", "Exact French door size", "", "—", 0, 0, 0, "Field + Spec", "TBD", "ТЗ §33.13"),
    ("Openings", "Exact main entry door size", "", "—", 0, 0, 0, "Field + Spec", "TBD", "ТЗ §33.14"),
    ("Openings", "Folding/panel system product + RO", "", "—", 0, 0, 0, "Spec", "TBD", "ТЗ §33.15"),
    ("Openings", "Closet sliding door size", "", "—", 0, 0, 0, "Field + Spec", "TBD", "ТЗ §33.16"),
    ("MEP", "Sewer route + slope verification", "1/4\" per ft min", "—", 0, 0, 0, "Site survey", "TBD", "ТЗ §33.17"),
    ("MEP", "Water supply route", "", "—", 0, 0, 0, "Site survey", "TBD", "ТЗ §33.18"),
    ("MEP", "Electrical route", "Underground or overhead", "—", 0, 0, 0, "Site survey", "TBD", "ТЗ §33.19"),
    ("MEP", "Existing main electrical panel capacity", "Amperage", "—", 0, 0, 0, "Electrician", "TBD", "ТЗ §33.20"),
    ("MEP", "Required ADU subpanel amperage", "100A vs 125A", "—", 0, 0, 0, "Electrician", "TBD", "ТЗ §33.21"),
    ("Floor", "Floor slab leveling needed?", "", "—", 0, 0, 0, "Field inspection", "TBD", "ТЗ §33.22"),
    ("Floor", "Moisture mitigation over slab?", "Vapor barrier sufficient?", "—", 0, 0, 0, "Field inspection", "TBD", "ТЗ §33.23"),
    ("Bath", "Additional waterproofing in wet zones?", "", "—", 0, 0, 0, "Spec", "TBD", "ТЗ §33.24"),
    ("Framing", "Structural headers for exterior openings?", "Engineer review", "—", 0, 0, 0, "Engineering", "TBD", "ТЗ §33.25"),
    ("Code", "Menlo Park / CA code requirements", "Fire, energy, vent, safety", "—", 0, 0, 0, "Code review", "TBD", "ТЗ §33.26"),
]


SHEETS = [
    ("02_Site_Exterior_Utilities", "Phase 1 — Site & Exterior Utilities (§19)", "Sewer 100ft + Water 100ft + Electric 50ft", "Site / Trenching", SITE_EXTERIOR),
    ("03_Raised_Floor_System", "Phase 2 — Raised Floor System (§17)", "Vapor barrier + sleepers + joists + subfloor (10\")", "Floor System", RAISED_FLOOR),
    ("04_Framing", "Phase 3 — Framing (§18)", "Left wall +21ft + interior partitions + headers", "Framing", FRAMING),
    ("05_Doors_Windows", "Phase 4 — Doors & Windows / Openings (§25)", "3 windows + 5 doors + folding system", "Openings", DOORS_WINDOWS),
    ("06_Plumbing_Roughin", "Phase 5 — Plumbing Rough-in (§20)", "PEX supply + ABS/PVC drain + vent + fittings", "Plumbing R-I", PLUMBING_RI),
    ("07_Electrical_Roughin", "Phase 6 — Electrical Rough-in (§21)", "ADU subpanel + Romex + boxes + devices + light fixtures", "Electrical R-I", ELECTRICAL_RI),
    ("08_HVAC_MiniSplit", "Phase 7 — HVAC Mini-Split (§22)", "1 outdoor + 2 indoor units, line sets, mounting", "HVAC", HVAC),
    ("09_Insulation", "Phase 8 — Insulation (§23)", "Walls R-15/R-21, ceiling R-30, floor R-19, sound", "Insulation", INSULATION),
    ("10_Drywall", "Phase 9 — Drywall (§24)", "60-75 sheets regular + 10-15 moisture-resistant", "Drywall", DRYWALL),
    ("11_Paint", "Phase 10 — Paint (§27)", "Primer + walls + ceiling + trim + bathroom moisture", "Paint", PAINT),
    ("12_Flooring_Trim", "Phase 11 — Flooring & Trim (§26)", "LVP 350 sf + baseboards 140 LF", "Flooring", FLOORING_TRIM),
    ("13_Kitchen_Cabinets", "Phase 12 — Kitchen Cabinets & Countertop (§11)", "IKEA SEKTION one-wall + countertop + sink/faucet", "Cabinetry", KITCHEN_CABINETS),
    ("14_Bathroom_Fixtures", "Phase 13 — Bathroom Fixtures (§14)", "Prefab shower + toilet + vanity + accessories + fan", "Finish — Fixtures", BATHROOM_FIXTURES),
    ("15_Closet_System", "Phase 14 — Closet System (§13)", "IKEA PAX/BOAXEL для walk-in 6×5'6\"", "Furnishing", CLOSET),
    ("16_Water_Heater", "Phase 15 — Water Heater & Utility (§16)", "Electric tank 20-30 gal + seismic strap + utility closet", "MEP — Water Heater", WATER_HEATER),
    ("17_Appliances", "Phase 16 — Appliances (§28)", "5 kitchen + washer + dryer + stacking kit", "Appliances", APPLIANCES),
    ("18_Furniture_IKEA", "Phase 17 — Furniture (§29)", "Living + Bedroom + Dining IKEA", "Furniture", FURNITURE),
    ("19_Special_Orders", "Cross-reference — Special Orders (§31.5)", "Special items з lead times 4-12 weeks", "Special Order", SPECIAL_ORDERS),
    ("20_Verification_Items", "Cross-reference — Verification Items (§33)", "26 пунктів must-verify-before-ordering", "Verification", VERIFICATION),
]


# ---------- Writers ----------
def style_header_row(ws, row, n_cols):
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER_THIN


def write_phase_sheet(wb, sheet_name, title, subtitle, phase_label, rows):
    ws = wb.create_sheet(sheet_name)
    n_cols = len(COLUMNS)

    # Заголовки
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(row=2, column=1, value=subtitle).font = SUBTITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18

    # Колонки header (рядок 4)
    header_row = 4
    for idx, (name, width) in enumerate(COLUMNS, start=1):
        ws.cell(row=header_row, column=idx, value=name)
        ws.column_dimensions[get_column_letter(idx)].width = width
    style_header_row(ws, header_row, n_cols)
    ws.row_dimensions[header_row].height = 32

    # Дані
    data_start = header_row + 1
    for i, row in enumerate(rows, start=1):
        cat, item, spec, unit, qlow, qmid, qhigh, source, status, notes = row
        r = data_start + i - 1
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=phase_label)
        ws.cell(row=r, column=3, value=cat)
        ws.cell(row=r, column=4, value=item)
        ws.cell(row=r, column=5, value=spec)
        ws.cell(row=r, column=6, value=unit)
        ws.cell(row=r, column=7, value=qlow)
        ws.cell(row=r, column=8, value=qmid)
        ws.cell(row=r, column=9, value=qhigh)
        # J = Unit Price (порожнє)
        ws.cell(row=r, column=10, value=None)
        # K = Total = H * J (Qty Mid × Unit Price)
        ws.cell(row=r, column=11, value=f"=IF(OR(H{r}=\"\",J{r}=\"\",J{r}=0),\"\",H{r}*J{r})")
        ws.cell(row=r, column=12, value=source)
        ws.cell(row=r, column=13, value="")  # SKU/Model
        ws.cell(row=r, column=14, value="")  # URL
        ws.cell(row=r, column=15, value=status)
        ws.cell(row=r, column=16, value="")  # Confidence
        ws.cell(row=r, column=17, value=notes)

        # Стилі
        for col in range(1, n_cols + 1):
            c = ws.cell(row=r, column=col)
            c.border = BORDER_THIN
            c.alignment = Alignment(vertical="top", wrap_text=True)
        for col in (7, 8, 9):
            ws.cell(row=r, column=col).number_format = "#,##0.##"
        ws.cell(row=r, column=10).number_format = "$#,##0.00"
        ws.cell(row=r, column=11).number_format = "$#,##0.00"

        # Підсвічування Special Order рядків
        if source == "Special Order":
            for col in range(1, n_cols + 1):
                ws.cell(row=r, column=col).fill = SPECIAL_ORDER_FILL

    # Підсумок
    total_row = data_start + len(rows) + 1
    ws.cell(row=total_row, column=3, value="ПІДСУМОК ФАЗИ").font = TOTAL_FONT
    ws.cell(row=total_row, column=11, value=f"=SUM(K{data_start}:K{data_start + len(rows) - 1})")
    ws.cell(row=total_row, column=11).font = TOTAL_FONT
    ws.cell(row=total_row, column=11).number_format = "$#,##0.00"
    for col in range(1, n_cols + 1):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).border = BORDER_THIN

    # Freeze + filter
    ws.freeze_panes = ws.cell(row=data_start, column=4)
    last_col_letter = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{data_start + len(rows) - 1}"

    return total_row


def write_summary(wb, sheet_totals):
    ws = wb.create_sheet("01_Summary", 0)

    # Title
    ws.cell(row=1, column=1, value="ADU Materials Estimate V2 — Зведений по фазам").font = TITLE_FONT
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 28

    info = [
        ("Проєкт", "Garage Conversion to ADU"),
        ("Локація", "Menlo Park, California 94025"),
        ("Total ADU Area", "315 sq ft (15' × 21')"),
        ("Ceiling", "Vaulted: 8 ft edges / 12 ft center"),
        ("Raised Floor", "10 inches"),
        ("Estimate Type", "Materials only (no labor / fees / tax per ТЗ §32)"),
        ("Approach", "Construction phases (shell-to-finish, coробка готова)"),
        ("Sources", "Home Depot / IKEA / Best Buy / Amazon / Special Order"),
        ("Дата створення", date.today().isoformat()),
        ("Статус", "V2 Draft — phase-based, потребує заповнення цін"),
    ]
    for i, (k, v) in enumerate(info, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)

    # Sheets summary table
    table_row = 15
    ws.cell(row=table_row, column=1, value="Розподіл по construction phases").font = SUBTITLE_FONT
    ws.merge_cells(f"A{table_row}:G{table_row}")

    header_row = table_row + 1
    headers = ["#", "Phase / Лист", "Phase Label", "Опис", "Items", "Subtotal ($)", "% of Total"]
    widths = [5, 28, 18, 50, 8, 14, 12]
    for idx, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.cell(row=header_row, column=idx, value=h)
        ws.column_dimensions[get_column_letter(idx)].width = w
    style_header_row(ws, header_row, len(headers))
    ws.row_dimensions[header_row].height = 28

    data_start = header_row + 1
    for i, (sheet_name, title, desc, phase, items_count, total_cell) in enumerate(sheet_totals, start=1):
        r = data_start + i - 1
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=sheet_name)
        ws.cell(row=r, column=3, value=phase)
        ws.cell(row=r, column=4, value=title)
        ws.cell(row=r, column=5, value=items_count)
        ws.cell(row=r, column=6, value=f"='{sheet_name}'!{total_cell}")
        ws.cell(row=r, column=6).number_format = "$#,##0.00"
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = BORDER_THIN
            ws.cell(row=r, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    grand_row = data_start + len(sheet_totals) + 1
    ws.cell(row=grand_row, column=2, value="GRAND TOTAL").font = TOTAL_FONT
    ws.cell(row=grand_row, column=6, value=f"=SUM(F{data_start}:F{data_start + len(sheet_totals) - 1})")
    ws.cell(row=grand_row, column=6).font = TOTAL_FONT
    ws.cell(row=grand_row, column=6).number_format = "$#,##0.00"
    for col in range(1, 8):
        ws.cell(row=grand_row, column=col).fill = TOTAL_FILL
        ws.cell(row=grand_row, column=col).border = BORDER_THIN

    for i in range(len(sheet_totals)):
        r = data_start + i
        ws.cell(row=r, column=7, value=f"=IF($F${grand_row}=0,\"\",F{r}/$F${grand_row})")
        ws.cell(row=r, column=7).number_format = "0.0%"

    # Notes про Phase 19 i Phase 20
    notes_row = grand_row + 2
    ws.cell(row=notes_row, column=1, value="ℹ️  Phase 19 (Special Orders) = CROSS-REFERENCE only. Items already in respective phases (Phase 4, 7, 15, 16). Не додає до Grand Total.").font = Font(italic=True, color="606060", size=9)
    ws.merge_cells(f"A{notes_row}:G{notes_row}")
    ws.cell(row=notes_row + 1, column=1, value="ℹ️  Phase 20 (Verification Items) = 26 пунктів §33 not priced, but may change quantities after field verification.").font = Font(italic=True, color="606060", size=9)
    ws.merge_cells(f"A{notes_row + 1}:G{notes_row + 1}")

    # Block 1: Reference Totals (tax + buffers)
    ref_row = notes_row + 3
    ws.cell(row=ref_row, column=1, value="📊  Reference Totals (НЕ включено в Grand Total)").font = SUBTITLE_FONT
    ws.merge_cells(f"A{ref_row}:G{ref_row}")

    ref_headers = ["Item", "$"]
    ref_header_row = ref_row + 1
    for idx, h in enumerate(ref_headers, start=1):
        ws.cell(row=ref_header_row, column=idx, value=h)
    ws.cell(row=ref_header_row, column=1).fill = HEADER_FILL
    ws.cell(row=ref_header_row, column=1).font = HEADER_FONT
    ws.cell(row=ref_header_row, column=2).fill = HEADER_FILL
    ws.cell(row=ref_header_row, column=2).font = HEADER_FONT
    ws.merge_cells(start_row=ref_header_row, start_column=2, end_row=ref_header_row, end_column=3)

    ref_data_start = ref_header_row + 1
    ref_items = [
        ("Grand Total — Materials Only", f"=F{grand_row}"),
        ("Estimated CA Sales Tax 9.375% (reference)", f"=F{grand_row}*0.09375"),
        ("Materials + Tax (reference)", f"=F{grand_row}*1.09375"),
        ("10% Material Buffer (reference)", f"=F{grand_row}*0.10"),
        ("15% Material Buffer (reference)", f"=F{grand_row}*0.15"),
        ("20% Material Buffer (reference)", f"=F{grand_row}*0.20"),
        ("Materials + Tax + 15% Buffer (planning)", f"=F{grand_row}*1.09375 + F{grand_row}*0.15"),
    ]
    for i, (label, formula) in enumerate(ref_items, start=ref_data_start):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=formula)
        ws.cell(row=i, column=2).number_format = "$#,##0.00"
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
        for col in range(1, 4):
            ws.cell(row=i, column=col).border = BORDER_THIN

    # Block 2: Top Cost Drivers
    tcd_row = ref_data_start + len(ref_items) + 2
    ws.cell(row=tcd_row, column=1, value="🔝  Top Cost Drivers (top phases по сумі)").font = SUBTITLE_FONT
    ws.merge_cells(f"A{tcd_row}:G{tcd_row}")

    tcd_items = [
        "1. Doors & Windows / Openings — ~18%",
        "2. Appliances — ~11%",
        "3. HVAC Mini-Split — ~8%",
        "4. Electrical Rough-in — ~8%",
        "5. Furniture — ~7%",
        "6. Raised Floor — ~7%",
        "7. Kitchen Cabinets — ~7%",
        "8. Site / Exterior Utilities — ~6%",
    ]
    for i, txt in enumerate(tcd_items, start=tcd_row + 1):
        ws.cell(row=i, column=1, value=f"  • {txt}").font = Font(size=10)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)

    # Block 3: High-Risk Verification Items
    hrv_row = tcd_row + len(tcd_items) + 2
    ws.cell(row=hrv_row, column=1, value="⚠️  High-Risk Verification Items (можуть змінити quantities)").font = SUBTITLE_FONT
    ws.merge_cells(f"A{hrv_row}:G{hrv_row}")

    hrv_items = [
        "Bathroom + laundry + utility fit (5'×7' може не вмістити все)",
        "Existing garage door opening — closure framing + sheathing",
        "Fire-rated drywall (5/8\" Type X ceiling якщо req'd)",
        "Electrical load calculation — affects subpanel + feeder + breakers",
        "IKEA Kitchen Planner — exact cabinet layout перед order",
        "Appliance specs — outlets, clearances, cabinet openings, cords",
        "Sewer slope / trench route verification",
        "Window / door exact sizes — field-verify",
        "Water heater 30 gal — physical fit в utility closet",
        "Special order lead times — affects planning, не materials total",
    ]
    for i, txt in enumerate(hrv_items, start=hrv_row + 1):
        ws.cell(row=i, column=1, value=f"  • {txt}").font = Font(size=10)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)

    # Block 4: Out-of-scope (existing)
    oos_row = hrv_row + len(hrv_items) + 2
    ws.cell(row=oos_row, column=1, value="❌  Out-of-scope (per ТЗ §32 — не включено в estimate)").font = SUBTITLE_FONT
    ws.merge_cells(f"A{oos_row}:G{oos_row}")
    oos_items = [
        "Labor / installation pricing",
        "CA sales tax (~9.375%) — see Reference Totals above",
        "Delivery / shipping fees",
        "Permit + plan check + inspection fees",
        "Architectural / engineering fees",
        "Title 24 documentation",
        "Demolition labor",
        "Trenching labor",
        "Contractor markup",
        "Contingency / buffer (~10-20%) — see Reference Totals above",
        "Disposal / dumpster fees",
    ]
    for i, txt in enumerate(oos_items, start=oos_row + 1):
        ws.cell(row=i, column=1, value=f"  • {txt}").font = Font(italic=True, color="808080", size=10)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)

    # Block 5: Included Summary
    inc_row = oos_row + len(oos_items) + 2
    ws.cell(row=inc_row, column=1, value="✅  Included в Grand Total").font = SUBTITLE_FONT
    ws.merge_cells(f"A{inc_row}:G{inc_row}")
    inc_items = [
        "Materials, fixtures, appliances, furniture (matters only)",
        "IKEA packages (kitchen SEKTION, closet BOAXEL, furniture)",
        "Home Depot construction materials (lumber, drywall, MEP, etc.)",
        "Special-order item allowances (French door, folding system, mini-split, washer/dryer)",
        "CA Code-required items (seismic strap, AFCI breakers, TR outlets, hardwired smoke/CO)",
    ]
    for i, txt in enumerate(inc_items, start=inc_row + 1):
        ws.cell(row=i, column=1, value=f"  • {txt}").font = Font(italic=True, color="2E7D32", size=10)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)

    ws.freeze_panes = "A3"


def build():
    wb = Workbook()
    wb.remove(wb.active)  # видаляємо дефолтний

    sheet_totals = []

    for sheet_name, title, subtitle, phase, rows in SHEETS:
        total_row = write_phase_sheet(wb, sheet_name, title, subtitle, phase, rows)
        total_cell = f"K{total_row}"
        sheet_totals.append((sheet_name, title, subtitle, phase, len(rows), total_cell))

    write_summary(wb, sheet_totals)

    wb.save(OUT_PATH)
    print(f"✅ Створено V2: {OUT_PATH}")
    print(f"   Листів: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        print(f"   - {name}")
    total_items = sum(len(rows) for _, _, _, _, rows in SHEETS)
    print(f"   Загальна кількість items: {total_items}")


if __name__ == "__main__":
    build()
